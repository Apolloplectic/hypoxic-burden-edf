"""
Hypoxic Burden Calculator - Polysomnography Analysis Tool
Based on: Azarbarzin A, et al. European Heart Journal (2019)
DOI: 10.1093/eurheartj/ehy624

Author: Sam Johnson
Email: sam.johnson9797@gmail.com
GitHub: https://github.com/Apolloplectic/hypoxic-burden-edf

UI/UX Redesign — tab-based layout with progressive disclosure.
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import os
from datetime import datetime
import io
import zipfile

# Import custom modules
from analysis_engine import PSGAnalyzer

try:
    from pdf_generator import PDFReportGenerator
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from utils import initialize_session_state, load_edf_file
from config import YASA_AVAILABLE, RISK_THRESHOLDS, get_risk_level
from validation import PSGValidator, check_zero_events, check_unrealistic_hb
from persistence import (
    save_analysis_to_session,
    display_analysis_history,
    compare_with_previous,
    display_comparison,
)

# =============================================================
# PAGE CONFIGURATION
# =============================================================
st.set_page_config(
    page_title="Hypoxic Burden Calculator",
    layout="centered",
    initial_sidebar_state="collapsed",
)

initialize_session_state()

# =============================================================
# PRESETS (shared across all tabs)
# =============================================================
PRESETS = {
    "Azarbarzin 2019 (Default)": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'azarbarzin',
        'description': "Exact parameters from Azarbarzin et al. EHJ 2019",
    },
    "Parekh 2023 (Automated)": {
        'pre_event_sec': 0,
        'desat_start_sec': 0,
        'desat_end_sec': 0,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'parekh',
        'description': "Parekh et al. AJRCCM 2023 — automated SpO\u2082 nadir detection with peak-prominence baselines",
    },
    "AASM 2023 Standard": {
        'pre_event_sec': 120,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 3,
        'artifact_filter': 'Mild (10%/s)',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'azarbarzin',
        'description': "Current clinical practice guidelines",
    },
    "Conservative (High Specificity)": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 4,
        'artifact_filter': 'Strict (5%/s)',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'azarbarzin',
        'description': "Minimizes false positives (4% threshold)",
    },
    "Aggressive (High Sensitivity)": {
        'pre_event_sec': 80,
        'desat_start_sec': 0,
        'desat_end_sec': 150,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'azarbarzin',
        'description': "Maximizes event detection",
    },
    "Custom": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'use_global_hb': True,
        'baseline_method': 'stage_specific',
        'floating_window_sec': 600,
        'method': 'azarbarzin',
        'description': "Manually configure all parameters",
    },
}


# =============================================================
# SHARED HELPER FUNCTIONS
# =============================================================

def _run_analysis(analyzer, params, use_mit_st=False):
    """Run analysis with given analyzer and params dict. Returns results dict."""
    if params.get('method') == 'parekh':
        return analyzer.run_parekh_analysis(
            artifact_filter=params['artifact_filter'],
            desat_threshold=params['desat_threshold'],
            use_global_hb=True,
            use_mit_st=use_mit_st,
            baseline_method=params.get('baseline_method', 'stage_specific'),
            floating_window_sec=params.get('floating_window_sec', 600),
        )
    return analyzer.run_full_analysis(
        pre_event_sec=params['pre_event_sec'],
        desat_start_sec=params['desat_start_sec'],
        desat_end_sec=params['desat_end_sec'],
        artifact_filter=params['artifact_filter'],
        desat_threshold=params['desat_threshold'],
        use_global_hb=True,
        preset_baseline=params['preset_baseline'],
        use_mit_st=use_mit_st,
        baseline_method=params.get('baseline_method', 'stage_specific'),
        floating_window_sec=params.get('floating_window_sec', 600),
    )


def _validate_and_detect(raw, analyzer, key_prefix=""):
    """
    Consolidated validation + channel detection.
    Auto-collapses when everything is OK.
    Returns True if analysis can proceed, False otherwise.
    """
    validator = PSGValidator(raw, analyzer)
    validation_results = validator.validate_all()

    all_ok = validation_results['valid'] and len(validator.warnings) == 0

    with st.expander(
        "✅ File Validation & Channels" if all_ok else "⚠️ File Validation & Channels",
        expanded=not all_ok,
    ):
        # Errors
        if validator.errors:
            for error in validator.errors:
                st.error(f"• {error}")

        # Warnings
        if validator.warnings:
            for warning in validator.warnings:
                st.warning(f"• {warning}")

        # Channel summary (compact)
        ch_col1, ch_col2, ch_col3 = st.columns(3)
        with ch_col1:
            st.write(f"**SpO\u2082:** {'✅ ' + analyzer.spo2_ch if analyzer.spo2_ch else '❌ Not found'}")
        with ch_col2:
            st.write(f"**Airflow:** {'✅ ' + analyzer.flow_ch if analyzer.flow_ch else '⚠️ Not found'}")
        with ch_col3:
            st.write(f"**EEG:** {'✅ ' + analyzer.eeg_ch if analyzer.eeg_ch else '⚠️ Not found'}")

        # Info items (collapsed within)
        if validator.info:
            with st.expander("ℹ️ Details", expanded=False):
                for info in validator.info:
                    st.info(f"• {info}")

        # Manual channel selection — only if auto-detection missed something
        needs_manual = not analyzer.spo2_ch or not analyzer.flow_ch or not analyzer.eeg_ch
        if needs_manual:
            st.markdown("---")
            st.markdown("**Manual Channel Selection**")
            all_channels = raw.ch_names
            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                if not analyzer.spo2_ch:
                    manual_spo2 = st.selectbox(
                        "SpO\u2082 Channel",
                        ["None"] + all_channels,
                        help="Required for analysis",
                        key=f"{key_prefix}manual_spo2",
                    )
                    if manual_spo2 != "None":
                        analyzer.spo2_ch = manual_spo2
            with mc2:
                if not analyzer.flow_ch:
                    manual_flow = st.selectbox(
                        "Airflow Channel",
                        ["None"] + all_channels,
                        help="Improves event detection",
                        key=f"{key_prefix}manual_flow",
                    )
                    if manual_flow != "None":
                        analyzer.flow_ch = manual_flow
            with mc3:
                if not analyzer.eeg_ch:
                    manual_eeg = st.selectbox(
                        "EEG Channel",
                        ["None"] + all_channels,
                        help="Enables sleep staging",
                        key=f"{key_prefix}manual_eeg",
                    )
                    if manual_eeg != "None":
                        analyzer.eeg_ch = manual_eeg

    # Stop if critical errors
    if not validation_results['valid']:
        st.error("**Cannot proceed — critical issues above must be resolved.**")
        return False
    if not analyzer.spo2_ch:
        st.error("❌ SpO\u2082 channel is required for analysis.")
        return False
    return True


def _settings_ui(key_prefix=""):
    """
    Unified settings UI: preset selector + single baseline method selector.
    Returns a params dict ready for _run_analysis().
    """
    st.subheader("⚙️ Analysis Settings")

    s_col1, s_col2 = st.columns([3, 1])
    with s_col1:
        preset_choice = st.selectbox(
            "Analysis Method",
            list(PRESETS.keys()),
            index=0,
            help="Choose methodology. Default (Azarbarzin 2019) works for most cases.",
            key=f"{key_prefix}preset_choice",
        )
    with s_col2:
        st.markdown("")  # spacer
        run_comparison = st.checkbox(
            "Compare presets",
            help="Run multiple presets side-by-side",
            key=f"{key_prefix}run_comparison",
        )

    comparison_presets = []
    if run_comparison:
        comparison_presets = st.multiselect(
            "Select presets to compare:",
            [p for p in PRESETS.keys() if p != "Custom"],
            default=["Azarbarzin 2019 (Default)", "AASM 2023 Standard"],
            help="Choose 2-4 presets",
            key=f"{key_prefix}comparison_presets",
        )
        if len(comparison_presets) < 2:
            st.warning("Select at least 2 presets to compare.")

    params = PRESETS[preset_choice].copy()

    # Advanced options — collapsed
    with st.expander("🔧 Advanced Options (optional)", expanded=False):
        st.caption(f"**{preset_choice}:** {params['description']}")

        if preset_choice == "Custom":
            st.markdown("**Event Detection**")
            c1, c2 = st.columns(2)
            with c1:
                params['pre_event_sec'] = st.slider(
                    "Pre-event baseline window (s)", 30, 180,
                    params['pre_event_sec'], key=f"{key_prefix}pre_event",
                )
                params['desat_start_sec'] = st.slider(
                    "Desaturation start offset (s)", 0, 120,
                    params['desat_start_sec'], key=f"{key_prefix}desat_start",
                )
            with c2:
                params['desat_end_sec'] = st.slider(
                    "Desaturation end (s after event end)", 30, 240,
                    params['desat_end_sec'], key=f"{key_prefix}desat_end",
                )
                params['artifact_filter'] = st.selectbox(
                    "Artifact filter",
                    ["Off", "Mild (10%/s)", "Strict (5%/s)"],
                    index=["Off", "Mild (10%/s)", "Strict (5%/s)"].index(params['artifact_filter']),
                    key=f"{key_prefix}artifact",
                )

            st.markdown("**Scoring**")
            sc1, sc2 = st.columns(2)
            with sc1:
                scoring = st.selectbox(
                    "Desaturation threshold",
                    ["3% (AASM)", "4% (Legacy)"],
                    index=0 if params['desat_threshold'] == 3 else 1,
                    key=f"{key_prefix}desat_thresh",
                )
                params['desat_threshold'] = 3 if "3%" in scoring else 4

        elif params.get('method') == 'parekh':
            # Show Parekh read-only info
            pc1, pc2 = st.columns(2)
            with pc1:
                st.write("**Method:** Automated SpO\u2082 nadir detection")
                st.write(f"**Nadir prominence:** ≥{params['desat_threshold']}% drop")
            with pc2:
                st.write("**Smoothing:** Savitzky-Golay (11 s)")
                st.write(f"**Artifact filter:** {params['artifact_filter']}")
        else:
            # Azarbarzin-family read-only
            pc1, pc2 = st.columns(2)
            with pc1:
                st.write(f"**Pre-event baseline:** {params['pre_event_sec']} s")
                st.write(f"**Desat start:** {params['desat_start_sec']} s from START")
            with pc2:
                st.write(f"**Desat end:** {params['desat_end_sec']} s after END")
                st.write(f"**Threshold:** {params['desat_threshold']}%")

        # --- Unified baseline method selector (one copy, not three) ---
        st.markdown("---")
        st.markdown("**Global HB Baseline Method**")
        baseline_options = [
            "Stage-specific (event-free epochs)",
            "Floating (trailing window)",
            "Whole-night (single value)",
        ]
        if preset_choice == "Custom":
            baseline_options.append("Manual entry")

        bl_choice = st.radio(
            "Baseline method",
            baseline_options,
            help="How to determine the baseline SpO\u2082 for global burden calculation",
            key=f"{key_prefix}bl_method",
            label_visibility="collapsed",
        )

        if bl_choice == "Manual entry":
            params['preset_baseline'] = st.slider(
                "Baseline SpO\u2082 (%)", 80.0, 100.0, 95.0, 0.1,
                format="%.1f", key=f"{key_prefix}manual_bl",
            )
            params['baseline_method'] = 'stage_specific'
        elif bl_choice.startswith("Floating"):
            params['baseline_method'] = 'floating'
            params['preset_baseline'] = 0.0
            params['floating_window_sec'] = st.slider(
                "Trailing window (minutes)", 2, 30, 10,
                help="Bottom 25% excluded, then 95th percentile.",
                key=f"{key_prefix}float_win",
            ) * 60
        elif bl_choice.startswith("Whole"):
            params['baseline_method'] = 'whole_night'
            params['preset_baseline'] = 0.0
        else:
            params['baseline_method'] = 'stage_specific'
            params['preset_baseline'] = 0.0

    return params, run_comparison, comparison_presets


def _display_results_summary(results, params):
    """Top-level results summary card (always visible)."""
    st.success("✅ Analysis Complete")
    st.markdown("### 📊 Results Summary")

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("AHI", f"{results['ahi']:.1f}", help="Apnea-Hypopnea Index (events/h)")
    with m2:
        st.metric(
            f"ODI ({params['desat_threshold']}%)",
            f"{results['odi']:.1f}",
            help=f"Oxygen Desaturation Index (≥{params['desat_threshold']}% drops/h)",
        )
    with m3:
        if len(results['events']) > 0:
            ci_str = f"[{results['ci'][0]:.1f}–{results['ci'][1]:.1f}]"
            st.metric("Obstructive HB", f"{results['total_hb']:.1f}", help=f"95% CI: {ci_str}")
            st.caption(f"95% CI: {ci_str}")
        else:
            st.metric("Obstructive HB", "0.0")

    # Risk level — prominent
    hb = results['total_hb']
    if hb >= RISK_THRESHOLDS['very_high']:
        risk_level, risk_emoji = "Very High", "🔴"
    elif hb >= RISK_THRESHOLDS['high']:
        risk_level, risk_emoji = "High", "🟠"
    elif hb >= RISK_THRESHOLDS['moderate']:
        risk_level, risk_emoji = "Moderate", "🟡"
    else:
        risk_level, risk_emoji = "Low", "🟢"

    # Progress-bar style risk indicator
    risk_pct = min(hb / 120.0, 1.0)  # cap at 120 for visual
    st.markdown(f"### {risk_emoji} Risk Level: **{risk_level}**")
    st.progress(risk_pct)

    # Method badge
    if results.get('method') == 'parekh':
        st.info(
            f"🔬 **Method: Parekh 2023** — {results.get('parekh_events_count', 0)} "
            "desaturation events detected via SpO\u2082 peak prominence."
        )

    # Sanity checks
    check_zero_events(results)
    check_unrealistic_hb(results)


def _display_results_details(results, params, analyzer, key_prefix=""):
    """Collapsible detailed result sections."""
    desat_threshold = params['desat_threshold']

    # --- SpO₂ trace ---
    if analyzer.df_spo2 is not None:
        with st.expander("📈 SpO\u2082 Signal & Events", expanded=False):
            spo2_df = analyzer.df_spo2.copy()
            spo2_times_h = spo2_df['time'] / 3600

            fig_spo2 = go.Figure()
            fig_spo2.add_trace(go.Scattergl(
                x=spo2_times_h, y=spo2_df['spo2'],
                mode='lines', name='SpO\u2082',
                line=dict(color='#1f77b4', width=1),
                hovertemplate='Time: %{x:.2f}h<br>SpO\u2082: %{y:.1f}%<extra></extra>',
            ))

            if results.get('floating_baseline') is not None:
                fb_arr = results['floating_baseline']
                if len(fb_arr) == len(spo2_times_h):
                    fig_spo2.add_trace(go.Scattergl(
                        x=spo2_times_h, y=fb_arr,
                        mode='lines', name='Floating Baseline',
                        line=dict(color='green', width=1.5, dash='dash'),
                    ))
            elif results.get('baseline_used'):
                fig_spo2.add_hline(
                    y=results['baseline_used'], line_dash="dash", line_color="green",
                    annotation_text=f"Baseline: {results['baseline_used']:.1f}%",
                    annotation_position="top left",
                )

            if results.get('baseline_used') and results.get('floating_baseline') is None:
                fig_spo2.add_hline(
                    y=results['baseline_used'] - desat_threshold,
                    line_dash="dot", line_color="orange",
                    annotation_text=f"{desat_threshold}% threshold",
                    annotation_position="bottom left", opacity=0.5,
                )

            fig_spo2.update_layout(
                xaxis_title="Time (hours)", yaxis_title="SpO\u2082 (%)",
                yaxis=dict(range=[60, 102]), height=400,
                margin=dict(l=50, r=20, t=30, b=50),
                hovermode='x unified', template='plotly_white',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            )
            fig_spo2.update_xaxes(rangeslider_visible=True, rangeslider_thickness=0.08)
            st.plotly_chart(fig_spo2, use_container_width=True)

    # --- Global HB ---
    if results.get('global_hb') is not None:
        with st.expander("🌍 Global Hypoxic Burden", expanded=False):
            bl_method = results.get('baseline_method', 'stage_specific')
            bl_method_labels = {
                'stage_specific': 'stage-specific baselines',
                'floating': f"floating baseline ({results.get('floating_window_sec', 600) // 60} min)",
                'whole_night': 'whole-night baseline',
            }

            g1, g2 = st.columns(2)
            with g1:
                st.metric(
                    "Global HB",
                    f"{results['global_hb']:.2f} (%min)/h",
                    help=f"Using {bl_method_labels.get(bl_method, bl_method)}",
                )
            with g2:
                if bl_method == 'floating' and results.get('floating_baseline') is not None:
                    st.metric("Mean Floating Baseline", f"{np.nanmean(results['floating_baseline']):.1f}%")
                elif results.get('baseline_used'):
                    st.metric("Baseline SpO\u2082", f"{results['baseline_used']:.1f}%")

            # Baseline detail sub-sections
            if bl_method == 'floating' and results.get('floating_baseline') is not None:
                fb = results['floating_baseline']
                fb1, fb2, fb3 = st.columns(3)
                with fb1:
                    st.metric("Mean", f"{np.nanmean(fb):.1f}%")
                with fb2:
                    st.metric("Min", f"{np.nanmin(fb):.1f}%")
                with fb3:
                    st.metric("Max", f"{np.nanmax(fb):.1f}%")
            elif bl_method == 'stage_specific':
                stage_baselines = results.get('stage_baselines', {})
                if stage_baselines:
                    fallback_bl = results.get('fallback_baseline', 0)
                    bl_data = []
                    for stg in ['W', 'N1', 'N2', 'N3', 'REM']:
                        if stg in stage_baselines:
                            bl_val = stage_baselines[stg]
                            src = "Event-free epochs" if abs(bl_val - fallback_bl) > 0.01 else "Fallback"
                            bl_data.append({'Stage': stg, 'Baseline SpO\u2082 (%)': f"{bl_val:.1f}", 'Source': src})
                    if bl_data:
                        st.dataframe(pd.DataFrame(bl_data), use_container_width=True, hide_index=True)

    # --- Signal quality ---
    if analyzer.df_spo2 is not None:
        with st.expander("📊 Signal Quality", expanded=False):
            spo2_vals = analyzer.df_spo2['spo2'].values
            q1, q2, q3 = st.columns(3)
            with q1:
                st.metric("Mean SpO\u2082", f"{np.nanmean(spo2_vals):.1f}%")
                st.metric("Median SpO\u2082", f"{np.nanmedian(spo2_vals):.1f}%")
            with q2:
                st.metric("Min SpO\u2082", f"{np.nanmin(spo2_vals):.1f}%")
                st.metric("Std Dev", f"{np.nanstd(spo2_vals):.2f}%")
            with q3:
                st.metric("Time <90%", f"{np.sum(spo2_vals < 90) / len(spo2_vals) * 100:.1f}%")
                st.metric("Time <88%", f"{np.sum(spo2_vals < 88) / len(spo2_vals) * 100:.1f}%")

    # --- Stage-specific ---
    if results.get('stage_hb'):
        with st.expander("😴 Sleep Stage Metrics", expanded=False):
            stage_rows = []
            for stg in ['W', 'N1', 'N2', 'N3', 'REM']:
                if stg in results['stage_hb']:
                    d = results['stage_hb'][stg]
                    stage_rows.append({
                        'Stage': stg,
                        'Time (h)': f"{d['hrs']:.1f}",
                        'AHI': f"{d['AHI']:.1f}",
                        'ODI': f"{d['ODI']:.1f}",
                        'HB': f"{d['HB']:.2f}",
                    })
            if stage_rows:
                st.dataframe(pd.DataFrame(stage_rows), use_container_width=True, hide_index=True)


def _export_section(results, params, filename, key_prefix=""):
    """Consolidated export buttons: PDF + Excel in one section."""
    st.markdown("---")
    st.markdown("### 📥 Export Reports")

    ex1, ex2 = st.columns(2)

    with ex1:
        if PDF_AVAILABLE:
            proof_mode = st.selectbox(
                "Proof plots",
                ["None", "Overlay (Azarbarzin-style)", "Full (all events)"],
                index=1,
                key=f"{key_prefix}proof_mode",
            )
            include_stages = st.checkbox("Include stages", value=True, key=f"{key_prefix}inc_stages")
            if st.button("📄 Generate PDF", type="primary", use_container_width=True, key=f"{key_prefix}pdf_btn"):
                with st.spinner("Generating PDF..."):
                    pdf_gen = PDFReportGenerator()
                    buf = pdf_gen.generate_report(
                        filename=filename, results=results,
                        proof_mode=proof_mode, include_stages=include_stages,
                    )
                    st.download_button(
                        "⬇️ Download PDF", buf.getvalue(),
                        file_name=f"HB_Report_{filename.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf", use_container_width=True,
                        key=f"{key_prefix}pdf_dl",
                    )
        else:
            st.info("PDF export unavailable (install reportlab).")

    with ex2:
        if st.button("📊 Generate Excel", use_container_width=True, key=f"{key_prefix}xlsx_btn"):
            with st.spinner("Generating Excel..."):
                buf = _build_excel_single(results, params, filename)
                st.download_button(
                    "⬇️ Download Excel", buf.getvalue(),
                    file_name=f"HB_Analysis_{filename.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"{key_prefix}xlsx_dl",
                )


def _build_excel_single(results, params, filename):
    """Build single-file Excel workbook and return BytesIO buffer."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "Hypoxic Burden Analysis Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = "File"; ws['B3'] = filename
    ws['A4'] = "Date"; ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A5'] = "Duration (h)"; ws['B5'] = results['duration']

    for col in ['A7', 'B7', 'C7']:
        ws[col].fill = hfill; ws[col].font = hfont
    ws['A7'] = "Metric"; ws['B7'] = "Value"; ws['C7'] = "95% CI"
    ws['A8'] = "AHI"; ws['B8'] = results['ahi']
    ws['A9'] = "ODI"; ws['B9'] = results['odi']
    ws['A10'] = "Obstructive HB"; ws['B10'] = results['total_hb']
    ws['C10'] = f"[{results['ci'][0]:.1f} - {results['ci'][1]:.1f}]"
    if results.get('global_hb'):
        ws['A11'] = "Global HB"; ws['B11'] = results['global_hb']
        ws['A12'] = "Baseline SpO\u2082"; ws['B12'] = results['baseline_used']
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

    # Stage sheet
    if results.get('stage_hb'):
        ws2 = wb.create_sheet("Stage-Specific")
        stage_data = []
        for stg in ['W', 'N1', 'N2', 'N3', 'REM']:
            if stg in results['stage_hb']:
                d = results['stage_hb'][stg]
                stage_data.append({'Stage': stg, 'Time (h)': d['hrs'], 'AHI': d['AHI'], 'ODI': d['ODI'], 'HB': d['HB']})
        df_s = pd.DataFrame(stage_data)
        for ri, row in enumerate(dataframe_to_rows(df_s, index=False, header=True), 1):
            for ci, val in enumerate(row, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                if ri == 1:
                    cell.fill = hfill; cell.font = hfont

    # Parameters sheet
    ws3 = wb.create_sheet("Parameters")
    ws3['A1'] = "Analysis Parameters"; ws3['A1'].font = Font(bold=True, size=14)
    ws3['A3'] = "Parameter"; ws3['B3'] = "Value"
    ws3['A3'].fill = hfill; ws3['B3'].fill = hfill
    ws3['A3'].font = hfont; ws3['B3'].font = hfont
    param_rows = [
        ("Pre-event baseline (s)", params['pre_event_sec']),
        ("Desat start (s)", params['desat_start_sec']),
        ("Desat end (s)", params['desat_end_sec']),
        ("Desat threshold (%)", params['desat_threshold']),
        ("Artifact filter", params['artifact_filter']),
        ("Global HB", "Yes" if params['use_global_hb'] else "No"),
    ]
    for i, (pname, pval) in enumerate(param_rows, 4):
        ws3[f'A{i}'] = pname; ws3[f'B{i}'] = pval
    ws3.column_dimensions['A'].width = 25; ws3.column_dimensions['B'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _display_comparison_analysis(comparison_results, comparison_data_rows, filename):
    """Display preset comparison table, charts, and export."""
    st.markdown("#### 📊 Comparison Table")
    df_cmp = pd.DataFrame(comparison_data_rows)
    st.dataframe(df_cmp, use_container_width=True, hide_index=True)

    # Charts
    st.markdown("#### 📈 Visual Comparison")
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
    presets_list = [d['Preset'] for d in comparison_data_rows]
    ahis = [float(d['AHI']) for d in comparison_data_rows]
    hbs = [float(d['Obstructive HB']) for d in comparison_data_rows]
    colors = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000']

    ax1.bar(range(len(presets_list)), ahis, color=colors[:len(presets_list)])
    ax1.set_xticks(range(len(presets_list)))
    ax1.set_xticklabels([p.split('(')[0].strip() for p in presets_list], rotation=15, ha='right')
    ax1.set_ylabel('AHI (events/hour)'); ax1.set_title('AHI by Preset'); ax1.grid(axis='y', alpha=0.3)

    ax2.bar(range(len(presets_list)), hbs, color=colors[:len(presets_list)])
    ax2.set_xticks(range(len(presets_list)))
    ax2.set_xticklabels([p.split('(')[0].strip() for p in presets_list], rotation=15, ha='right')
    ax2.set_ylabel('HB (%min/h)'); ax2.set_title('Obstructive HB by Preset'); ax2.grid(axis='y', alpha=0.3)

    plt.tight_layout()
    st.pyplot(fig)
    plt.close(fig)

    # Insights
    ahi_range = max(ahis) - min(ahis)
    hb_range = max(hbs) - min(hbs)
    i1, i2 = st.columns(2)
    with i1:
        st.metric("AHI Range", f"{ahi_range:.1f}")
    with i2:
        st.metric("HB Range", f"{hb_range:.1f}")
    if ahi_range > 5:
        st.warning(f"Wide AHI variation ({ahi_range:.1f}) — methodology significantly affects detection.")
    else:
        st.success(f"Consistent AHI across presets ({ahi_range:.1f} range).")

    # Comparison Excel export
    if st.button("📥 Download Comparison Excel", use_container_width=True, key="cmp_xlsx"):
        with st.spinner("Generating comparison workbook..."):
            buf = _build_excel_comparison(comparison_results, comparison_data_rows, filename)
            st.download_button(
                "⬇️ Download", buf.getvalue(),
                file_name=f"HB_Comparison_{filename.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="cmp_xlsx_dl",
            )


def _build_excel_comparison(comparison_results, comparison_data_rows, filename):
    """Build comparison Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")

    ws = wb.active; ws.title = "Comparison Summary"
    ws['A1'] = "Preset Comparison Analysis"; ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"File: {filename}"
    ws['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = ['Preset', 'AHI', 'ODI', 'Obstructive HB', 'CI Lower', 'CI Upper', 'Risk']
    for ci, h in enumerate(headers):
        cell = ws.cell(row=5, column=ci + 1, value=h)
        cell.fill = hfill; cell.font = hfont
    for ri, data in enumerate(comparison_data_rows, 6):
        ws.cell(row=ri, column=1, value=data['Preset'])
        ws.cell(row=ri, column=2, value=float(data['AHI']))
        ws.cell(row=ri, column=3, value=float(data['ODI']))
        ws.cell(row=ri, column=4, value=float(data['Obstructive HB']))
        ws.cell(row=ri, column=5, value=float(data['CI Lower']))
        ws.cell(row=ri, column=6, value=float(data['CI Upper']))
        ws.cell(row=ri, column=7, value=data['Risk'])
    ws.column_dimensions['A'].width = 30

    # Per-preset detail sheets
    for pname, presults in comparison_results.items():
        sname = pname.replace(" (Default)", "").replace(" (High Specificity)", "").replace(" (High Sensitivity)", "")[:31]
        ws2 = wb.create_sheet(sname)
        ws2['A1'] = f"{pname} — Detailed"; ws2['A1'].font = Font(bold=True, size=12)
        for ci, h in enumerate(["Metric", "Value"], 1):
            c = ws2.cell(row=3, column=ci, value=h); c.fill = hfill; c.font = hfont
        rows = [
            ("Duration (h)", presults['duration']),
            ("AHI", presults['ahi']),
            ("ODI", presults['odi']),
            ("Obstructive HB", presults['total_hb']),
            ("95% CI Lower", presults['ci'][0]),
            ("95% CI Upper", presults['ci'][1]),
        ]
        if presults.get('global_hb'):
            rows.append(("Global HB", presults['global_hb']))
            rows.append(("Baseline SpO\u2082", presults['baseline_used']))
        for ri, (metric, val) in enumerate(rows, 4):
            ws2.cell(row=ri, column=1, value=metric)
            ws2.cell(row=ri, column=2, value=val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# =============================================================
# HEADER
# =============================================================
st.title("🫁 Hypoxic Burden Calculator")
st.caption(
    "Upload PSG EDF file → get comprehensive sleep apnea metrics with 95% CI.  \n"
    "Based on Azarbarzin et al. *European Heart Journal* (2019) — "
    "[DOI: 10.1093/eurheartj/ehy624](https://doi.org/10.1093/eurheartj/ehy624)"
)

# Sidebar — analysis history
display_analysis_history()

# =============================================================
# MODE SELECTOR (TABS)
# =============================================================
tab_single, tab_compare, tab_batch = st.tabs([
    "📄 Single File",
    "🔬 Treatment Comparison",
    "📦 Batch Mode",
])


# =============================================================
# TAB 1: SINGLE FILE ANALYSIS
# =============================================================
with tab_single:
    edf_file = st.file_uploader(
        "Upload PSG EDF file",
        type=["edf"],
        help="Up to 200 MB online. Run locally for larger files.",
        key="single_upload",
    )
    st.caption("📥 File too large? [Run locally](https://github.com/Apolloplectic/hypoxic-burden-edf#local-install) for 2 GB+ support.")

    if edf_file is not None:
        # --- Load ---
        with st.spinner(f"Loading {edf_file.name} ({edf_file.size / 1e6:.1f} MB)..."):
            raw, temp_path = load_edf_file(edf_file)
        if raw is None:
            st.error("❌ Failed to load EDF file.")
            st.stop()
        st.success(f"Loaded — Duration: {raw.times[-1]/3600:.2f} hours")

        analyzer = PSGAnalyzer(raw, temp_path)

        # --- Validation & Channels (consolidated) ---
        can_proceed = _validate_and_detect(raw, analyzer, key_prefix="single_")
        if not can_proceed:
            st.stop()

        # MIT annotations
        if analyzer.check_mit_annotations():
            use_mit = st.checkbox(
                "Use MIT Gold Standard Annotations",
                value=True,
                help="MIT-annotated sleep stages and events from SHHS/slpdb",
                key="single_mit",
            )
            st.session_state.use_mit_st = use_mit
            if use_mit:
                st.caption(f"MIT: {len(analyzer.manual_events)} events, AHI = {analyzer.manual_ahi:.1f}")
        else:
            st.session_state.use_mit_st = False

        # --- Settings (unified) ---
        params, run_comparison, comparison_presets = _settings_ui(key_prefix="single_")

        # --- Analyze button ---
        st.markdown("---")
        if '_single_analyzed' not in st.session_state:
            st.session_state._single_analyzed = False

        if not st.session_state._single_analyzed:
            btn_label = "🔬 Compare Presets" if run_comparison else "🚀 Analyze File"
            if st.button(btn_label, type="primary", use_container_width=True, key="single_go"):
                st.session_state._single_analyzed = True
                st.session_state._single_do_comparison = run_comparison
                if run_comparison:
                    st.session_state._single_cmp_presets = comparison_presets
                st.rerun()
        else:
            # ---- Comparison mode ----
            if st.session_state.get('_single_do_comparison') and len(st.session_state.get('_single_cmp_presets', [])) >= 2:
                st.subheader("🔬 Preset Comparison Results")
                comparison_results = {}
                comparison_data_rows = []
                with st.spinner("Running analysis with multiple presets..."):
                    prog = st.progress(0)
                    for idx, pname in enumerate(st.session_state._single_cmp_presets):
                        p = PRESETS[pname]
                        res = _run_analysis(analyzer, p, st.session_state.use_mit_st)
                        comparison_results[pname] = res
                        comparison_data_rows.append({
                            'Preset': pname,
                            'AHI': f"{res['ahi']:.1f}",
                            'ODI': f"{res['odi']:.1f}",
                            'Obstructive HB': f"{res['total_hb']:.1f}",
                            'CI Lower': f"{res['ci'][0]:.1f}",
                            'CI Upper': f"{res['ci'][1]:.1f}",
                            'Risk': get_risk_level(res['total_hb']),
                        })
                        prog.progress((idx + 1) / len(st.session_state._single_cmp_presets))
                    prog.empty()
                _display_comparison_analysis(comparison_results, comparison_data_rows, edf_file.name)
                # Use first preset for detailed view
                results = list(comparison_results.values())[0]
                params = PRESETS[list(comparison_results.keys())[0]]
                st.markdown("---")
                st.markdown("### Detailed Results (First Preset)")
            else:
                # ---- Single analysis ----
                with st.spinner("🔬 Analyzing PSG data..."):
                    results = _run_analysis(analyzer, params, st.session_state.use_mit_st)

            # Save to history
            save_analysis_to_session(edf_file.name, results, params)

            # Compare with previous
            previous = compare_with_previous()
            if previous:
                display_comparison(results, previous)
                st.markdown("---")

            # Results
            _display_results_summary(results, params)
            _display_results_details(results, params, analyzer, key_prefix="single_")
            _export_section(results, params, edf_file.name, key_prefix="single_")

            # Reset
            st.markdown("---")
            if st.button("🔄 Analyze Another File", use_container_width=True, key="single_reset"):
                st.session_state._single_analyzed = False
                st.session_state._single_do_comparison = False
                st.session_state._single_cmp_presets = []
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                st.rerun()


# =============================================================
# TAB 2: TREATMENT COMPARISON
# =============================================================
with tab_compare:
    st.info("Compare the same patient before and after treatment (CPAP, Inspire, surgery, etc.)")

    tc1, tc2 = st.columns(2)
    with tc1:
        st.markdown("**Pre-Treatment (Baseline)**")
        pre_file = st.file_uploader("Upload baseline PSG", type=["edf"], key="pre_upload")
    with tc2:
        st.markdown("**Post-Treatment**")
        post_file = st.file_uploader("Upload post-treatment PSG", type=["edf"], key="post_upload")

    if pre_file and post_file:
        # Load both
        with st.spinner("Loading pre-treatment PSG..."):
            raw_pre, tp_pre = load_edf_file(pre_file)
        with st.spinner("Loading post-treatment PSG..."):
            raw_post, tp_post = load_edf_file(post_file)

        if raw_pre is None or raw_post is None:
            st.error("❌ Failed to load one or both files.")
            st.stop()

        analyzer_pre = PSGAnalyzer(raw_pre, tp_pre)
        analyzer_post = PSGAnalyzer(raw_post, tp_post)

        # Validation (compact, side-by-side)
        v1, v2 = st.columns(2)
        with v1:
            st.markdown("**Pre-Treatment**")
            st.write(f"Duration: {raw_pre.times[-1]/3600:.2f} h")
            st.write(f"SpO\u2082: {'✅ ' + analyzer_pre.spo2_ch if analyzer_pre.spo2_ch else '❌'}")
        with v2:
            st.markdown("**Post-Treatment**")
            st.write(f"Duration: {raw_post.times[-1]/3600:.2f} h")
            st.write(f"SpO\u2082: {'✅ ' + analyzer_post.spo2_ch if analyzer_post.spo2_ch else '❌'}")

        if not analyzer_pre.spo2_ch or not analyzer_post.spo2_ch:
            st.error("SpO\u2082 channel not found in one or both files.")
            st.stop()

        # Settings
        st.markdown("---")
        tc_preset = st.selectbox(
            "Analysis Preset (applied to both files)",
            list(PRESETS.keys()), index=0,
            help="Use same preset for fair comparison",
            key="tc_preset",
        )
        tc_params = PRESETS[tc_preset].copy()
        st.caption(f"{tc_preset}: {tc_params['description']}")

        # Run button
        if st.button("🚀 Run Comparison", type="primary", use_container_width=True, key="tc_go"):
            with st.spinner("Analyzing pre-treatment..."):
                res_pre = _run_analysis(analyzer_pre, tc_params, False)
            with st.spinner("Analyzing post-treatment..."):
                res_post = _run_analysis(analyzer_post, tc_params, False)

            # --- Treatment Effect Summary ---
            st.markdown("---")
            st.markdown("## 📊 Treatment Effect Summary")

            # Metrics with deltas
            d1, d2, d3 = st.columns(3)
            with d1:
                ahi_delta = res_post['ahi'] - res_pre['ahi']
                st.metric("AHI", f"{res_post['ahi']:.1f}", f"{ahi_delta:+.1f}", delta_color="inverse")
                st.caption(f"Pre: {res_pre['ahi']:.1f}")
            with d2:
                odi_delta = res_post['odi'] - res_pre['odi']
                st.metric("ODI", f"{res_post['odi']:.1f}", f"{odi_delta:+.1f}", delta_color="inverse")
                st.caption(f"Pre: {res_pre['odi']:.1f}")
            with d3:
                hb_delta = res_post['total_hb'] - res_pre['total_hb']
                st.metric("HB", f"{res_post['total_hb']:.1f}", f"{hb_delta:+.1f}", delta_color="inverse")
                st.caption(f"Pre: {res_pre['total_hb']:.1f}")

            # Efficacy interpretation
            ahi_imp = ((res_pre['ahi'] - res_post['ahi']) / res_pre['ahi'] * 100) if res_pre['ahi'] > 0 else 0
            hb_imp = ((res_pre['total_hb'] - res_post['total_hb']) / res_pre['total_hb'] * 100) if res_pre['total_hb'] > 0 else 0

            if ahi_imp >= 50 and hb_imp >= 50:
                st.success(f"✅ Excellent Response — AHI {ahi_imp:.0f}% ↓, HB {hb_imp:.0f}% ↓")
            elif ahi_imp >= 30 and hb_imp >= 30:
                st.info(f"Good Response — AHI {ahi_imp:.0f}% ↓, HB {hb_imp:.0f}% ↓")
            elif ahi_imp >= 10 and hb_imp >= 10:
                st.warning(f"Partial Response — AHI {ahi_imp:.0f}% ↓, HB {hb_imp:.0f}% ↓")
            else:
                st.error(f"Poor Response — AHI {ahi_imp:.0f}% change, HB {hb_imp:.0f}% change")

            # Detailed table
            with st.expander("📋 Full Comparison Table", expanded=False):
                comp_df = pd.DataFrame({
                    'Metric': ['AHI', 'ODI', 'Event-Specific HB', 'Global HB'],
                    'Pre': [f"{res_pre['ahi']:.1f}", f"{res_pre['odi']:.1f}",
                            f"{res_pre['total_hb']:.1f}", f"{res_pre.get('global_hb', 0):.1f}"],
                    'Post': [f"{res_post['ahi']:.1f}", f"{res_post['odi']:.1f}",
                             f"{res_post['total_hb']:.1f}", f"{res_post.get('global_hb', 0):.1f}"],
                    'Change': [f"{res_post['ahi'] - res_pre['ahi']:+.1f}",
                               f"{res_post['odi'] - res_pre['odi']:+.1f}",
                               f"{res_post['total_hb'] - res_pre['total_hb']:+.1f}",
                               f"{res_post.get('global_hb', 0) - res_pre.get('global_hb', 0):+.1f}"],
                })
                st.dataframe(comp_df, use_container_width=True, hide_index=True)

            # Stage comparison
            with st.expander("😴 Stage-Specific Comparison", expanded=False):
                stage_rows = []
                for stg in ['REM', 'N3', 'N2', 'N1', 'W']:
                    if stg in res_pre.get('stage_hb', {}) and stg in res_post.get('stage_hb', {}):
                        pre_s = res_pre['stage_hb'][stg]
                        post_s = res_post['stage_hb'][stg]
                        hb_pct = ((pre_s['HB'] - post_s['HB']) / pre_s['HB'] * 100) if pre_s['HB'] > 0 else 0
                        stage_rows.append({
                            'Stage': stg,
                            'Pre AHI': f"{pre_s['AHI']:.1f}", 'Post AHI': f"{post_s['AHI']:.1f}",
                            'Pre HB': f"{pre_s['HB']:.1f}", 'Post HB': f"{post_s['HB']:.1f}",
                            'HB Improvement': f"{hb_pct:.0f}%",
                        })
                if stage_rows:
                    st.dataframe(pd.DataFrame(stage_rows), use_container_width=True, hide_index=True)

            # Charts
            with st.expander("📈 Visual Comparison", expanded=False):
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
                metrics = ['AHI', 'ODI']
                pre_vals = [res_pre['ahi'], res_pre['odi']]
                post_vals = [res_post['ahi'], res_post['odi']]
                x = np.arange(len(metrics)); w = 0.35
                ax1.bar(x - w/2, pre_vals, w, label='Pre', color='#e74c3c', alpha=0.8)
                ax1.bar(x + w/2, post_vals, w, label='Post', color='#27ae60', alpha=0.8)
                ax1.set_ylabel('Events/h'); ax1.set_title('AHI & ODI'); ax1.set_xticks(x)
                ax1.set_xticklabels(metrics); ax1.legend(); ax1.grid(axis='y', alpha=0.3)

                hb_m = ['Event HB', 'Global HB']
                hb_pre = [res_pre['total_hb'], res_pre.get('global_hb', 0)]
                hb_post = [res_post['total_hb'], res_post.get('global_hb', 0)]
                x2 = np.arange(len(hb_m))
                ax2.bar(x2 - w/2, hb_pre, w, label='Pre', color='#e74c3c', alpha=0.8)
                ax2.bar(x2 + w/2, hb_post, w, label='Post', color='#27ae60', alpha=0.8)
                ax2.set_ylabel('(%·min)/h'); ax2.set_title('Hypoxic Burden'); ax2.set_xticks(x2)
                ax2.set_xticklabels(hb_m); ax2.legend(); ax2.grid(axis='y', alpha=0.3)
                plt.tight_layout(); st.pyplot(fig); plt.close(fig)

            # Export
            st.markdown("---")
            st.markdown("### 📥 Export")
            if st.button("📊 Download Comparison Excel", use_container_width=True, key="tc_xlsx"):
                with st.spinner("Generating..."):
                    excel_buf = io.BytesIO()
                    with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
                        summary_df = pd.DataFrame({
                            'Metric': ['AHI', 'ODI', 'Event HB', 'Global HB', 'Duration (h)', 'Events'],
                            'Pre': [res_pre['ahi'], res_pre['odi'], res_pre['total_hb'],
                                    res_pre.get('global_hb', 0), res_pre['duration'], len(res_pre.get('events', []))],
                            'Post': [res_post['ahi'], res_post['odi'], res_post['total_hb'],
                                     res_post.get('global_hb', 0), res_post['duration'], len(res_post.get('events', []))],
                            'Change': [
                                res_post['ahi'] - res_pre['ahi'],
                                res_post['odi'] - res_pre['odi'],
                                res_post['total_hb'] - res_pre['total_hb'],
                                res_post.get('global_hb', 0) - res_pre.get('global_hb', 0),
                                res_post['duration'] - res_pre['duration'],
                                len(res_post.get('events', [])) - len(res_pre.get('events', [])),
                            ],
                        })
                        summary_df.to_excel(writer, sheet_name='Summary', index=False)

                        if stage_rows:
                            pd.DataFrame(stage_rows).to_excel(writer, sheet_name='Stage Comparison', index=False)

                        interp = pd.DataFrame({
                            'Assessment': ['Efficacy', 'AHI', 'HB'],
                            'Result': [
                                'Excellent' if (ahi_imp >= 50 and hb_imp >= 50) else
                                'Good' if (ahi_imp >= 30 and hb_imp >= 30) else
                                'Partial' if (ahi_imp >= 10 and hb_imp >= 10) else 'Poor',
                                f"{ahi_imp:.1f}% improvement",
                                f"{hb_imp:.1f}% improvement",
                            ],
                        })
                        interp.to_excel(writer, sheet_name='Interpretation', index=False)

                    excel_buf.seek(0)
                    st.download_button(
                        "⬇️ Download", excel_buf.getvalue(),
                        file_name=f"treatment_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="tc_xlsx_dl",
                    )


# =============================================================
# TAB 3: BATCH MODE
# =============================================================
with tab_batch:
    st.info("Analyze multiple files with the same settings. Online: up to 5 files / 1 GB total.")

    batch_files = st.file_uploader(
        "Upload PSG EDF files",
        type=["edf"],
        accept_multiple_files=True,
        key="batch_upload",
    )
    st.caption("📥 For >5 files or >1 GB, [run locally](https://github.com/Apolloplectic/hypoxic-burden-edf#local-install).")

    if batch_files:
        n_files = len(batch_files)
        total_mb = sum(f.size for f in batch_files) / 1e6

        if n_files > 5 or total_mb > 1024:
            st.error(f"Batch too large ({n_files} files, {total_mb:.0f} MB). Limit: 5 files / 1 GB online.")
            st.stop()

        st.write(f"**{n_files} files** — {total_mb:.0f} MB total")

        # Settings — reference single-file presets
        b_col1, b_col2 = st.columns(2)
        with b_col1:
            batch_preset = st.selectbox(
                "Preset", list(PRESETS.keys()), index=0, key="batch_preset",
            )
        with b_col2:
            batch_desat = st.selectbox("Threshold", ["3%", "4%"], key="batch_desat")

        batch_params = PRESETS[batch_preset].copy()
        batch_params['desat_threshold'] = 3 if "3%" in batch_desat else 4

        with st.expander("More options", expanded=False):
            batch_proof = st.selectbox(
                "Proof plots", ["None", "Overlay (Azarbarzin-style)", "Full"],
                index=1, key="batch_proof",
            )
            batch_stages = st.checkbox("Include stages", value=True, key="batch_stages")

        # Run batch (simple — no pause/stop)
        if st.button("🚀 Run Batch", type="primary", use_container_width=True, key="batch_go"):
            prog = st.progress(0)
            status = st.empty()
            batch_summary = []
            batch_pdfs = []

            for idx, bf in enumerate(batch_files):
                status.text(f"Processing {bf.name} ({idx+1}/{n_files})...")
                prog.progress((idx + 0.1) / n_files)

                try:
                    raw_b, tp_b = load_edf_file(bf)
                    if raw_b is None:
                        st.warning(f"Skipping {bf.name}: could not load.")
                        continue

                    anlzr = PSGAnalyzer(raw_b, tp_b)
                    res = _run_analysis(anlzr, batch_params, False)

                    row = {
                        'File': bf.name,
                        'Duration (h)': f"{res['duration']:.1f}",
                        'AHI': f"{res['ahi']:.1f}",
                        'ODI': f"{res['odi']:.1f}",
                        'HB': f"{res['total_hb']:.2f}",
                    }
                    if res.get('global_hb') is not None:
                        row['Global HB'] = f"{res['global_hb']:.2f}"
                    batch_summary.append(row)

                    if PDF_AVAILABLE:
                        pdf_gen = PDFReportGenerator()
                        pdf_buf = pdf_gen.generate_report(
                            filename=bf.name, results=res,
                            proof_mode=batch_proof, include_stages=batch_stages,
                        )
                        batch_pdfs.append((bf.name, pdf_buf))

                    if os.path.exists(tp_b):
                        os.remove(tp_b)

                except Exception as e:
                    st.error(f"Error on {bf.name}: {e}")

                prog.progress((idx + 1) / n_files)

            prog.progress(1.0)
            status.text("✅ Batch complete!")

            # Summary table
            if batch_summary:
                st.dataframe(pd.DataFrame(batch_summary), use_container_width=True, hide_index=True)

            # ZIP download
            if PDF_AVAILABLE and batch_pdfs:
                status.text("Packaging ZIP...")
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for fname, pbuf in batch_pdfs:
                        zf.writestr(f"Reports/HB_Report_{fname.replace('.edf', '')}.pdf", pbuf.getvalue())
                    # Master summary
                    pdf_gen = PDFReportGenerator()
                    master_buf = pdf_gen.generate_batch_summary(batch_summary)
                    zf.writestr("Master_Summary.pdf", master_buf.getvalue())
                zip_buf.seek(0)

                st.download_button(
                    f"⬇️ Download All Reports ({len(batch_pdfs)} files)",
                    zip_buf.getvalue(),
                    file_name=f"HB_Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip", type="primary", use_container_width=True,
                    key="batch_zip_dl",
                )
            status.empty()


# =============================================================
# LOCAL INSTRUCTIONS (collapsed, bottom)
# =============================================================
with st.expander("📥 Run locally for large files (2 GB+)", expanded=False):
    st.markdown("""
### How to Run Locally (No Coding Required)

**Step 1:** Download Python 3.9+ from [python.org](https://www.python.org/downloads/) — check "Add to PATH".

**Step 2:** Download the app from [GitHub](https://github.com/Apolloplectic/hypoxic-burden-edf/releases), unzip, open terminal in folder, run:
```
pip install -r requirements.txt
```

**Step 3:** Run:
```
streamlit run app.py --server.maxUploadSize=4096
```

Need help? [sam.johnson9797@gmail.com](mailto:sam.johnson9797@gmail.com) | [GitHub Issues](https://github.com/Apolloplectic/hypoxic-burden-edf/issues)
""")


# =============================================================
# FOOTER
# =============================================================
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666;'>"
    "<p><strong>Hypoxic Burden Calculator</strong> — Open Source Sleep Apnea Analysis</p>"
    "<p>"
    "🐙 <a href='https://github.com/Apolloplectic/hypoxic-burden-edf'>GitHub</a> • "
    "📄 <a href='https://doi.org/10.5281/zenodo.17561726'>DOI: 10.5281/zenodo.17561726</a> • "
    "📧 <a href='mailto:sam.johnson9797@gmail.com'>Contact</a>"
    "</p>"
    "<p><small>Built with Streamlit • MNE • {yasa} • WFDB</small></p>"
    "<p><small>Cite: Azarbarzin A, et al. <em>Eur Heart J</em> 2019;40:1149-1157</small></p>"
    "</div>".format(yasa="YASA ✅" if YASA_AVAILABLE else "YASA ❌"),
    unsafe_allow_html=True,
)

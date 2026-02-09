"""
Hypoxic Burden Calculator - Polysomnography Analysis Tool
Based on: Azarbarzin A, et al. European Heart Journal (2019)
DOI: 10.1093/eurheartj/ehy624

Author: Sam Johnson
Email: sam.johnson9797@gmail.com
GitHub: https://github.com/Apolloplectic/hypoxic-burden-edf
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from datetime import datetime
import io
import zipfile

# Import custom modules
from analysis_engine import PSGAnalyzer
try:
    from pdf_generator import PDFReportGenerator
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ pdf_generator.py not found - PDF export disabled")
    
from utils import initialize_session_state, load_edf_file
from config import YASA_AVAILABLE
from validation import PSGValidator, check_zero_events, check_unrealistic_hb
from persistence import save_analysis_to_session, display_analysis_history, compare_with_previous, display_comparison

# --------------------------------------------------------------
# PAGE CONFIGURATION
# --------------------------------------------------------------
st.set_page_config(
    page_title="Hypoxic Burden Calculator",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# --------------------------------------------------------------
# HEADER
# --------------------------------------------------------------
st.title("🫁 Hypoxic Burden Calculator")
st.markdown("""
**Upload PSG EDF file** → get comprehensive sleep apnea metrics with **95% CI**.

Based on:
> Azarbarzin A, et al. *European Heart Journal* (2019) – DOI: [10.1093/eurheartj/ehy624](https://doi.org/10.1093/eurheartj/ehy624)
""")

# --------------------------------------------------------------
# FILE UPLOAD SECTION
# --------------------------------------------------------------
st.markdown("### 📁 Upload Mode")

# Mode selection
analysis_mode = st.radio(
    "Select Analysis Mode:",
    ["Single File Analysis", "Treatment Comparison (Before/After) - Broken at the moment"],
    help="Compare pre-treatment vs post-treatment PSG studies"
)

if analysis_mode == "Single File Analysis":
    st.markdown("#### Upload PSG EDF file")
    edf_file = st.file_uploader(
        "Upload PSG EDF file",
        type=["edf"],
        help="⚠️ Online version limited to 200 MB. For larger files (up to 2 GB), run locally (see instructions below).",
        key="single_file_upload"
    )
    pre_treatment_file = None
    post_treatment_file = None

else:  # Treatment Comparison mode
    st.markdown("#### Upload Before & After PSG Files")
    st.info("📊 Compare the same patient before and after treatment (CPAP, Inspire, surgery, etc.)")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Pre-Treatment (Baseline)**")
        pre_treatment_file = st.file_uploader(
            "Upload baseline PSG",
            type=["edf"],
            help="PSG before any treatment",
            key="pre_treatment_upload"
        )
    
    with col2:
        st.markdown("**Post-Treatment**")
        post_treatment_file = st.file_uploader(
            "Upload post-treatment PSG",
            type=["edf"],
            help="PSG after treatment (CPAP, Inspire, etc.)",
            key="post_treatment_upload"
        )
    
    edf_file = None  # Not used in comparison mode

# --------------------------------------------------------------
# LOCAL RUN INSTRUCTIONS
# --------------------------------------------------------------
with st.expander("📥 File too large? Run locally (2 GB+ support) — no coding needed!", expanded=False):
    st.markdown("""
    ### **How to Run This App on Your Computer (2 GB+ Files)**
    **No programming experience required. Takes 5 minutes.**
    
    ---
    
    #### **Step 1: Install Python (if not already installed)**
    1. Download Python 3.9+ from [python.org](https://www.python.org/downloads/)
    2. During installation, **check "Add Python to PATH"**
    
    ---
    
    #### **Step 2: Download & Setup**
    1. Download the app from [GitHub Releases](https://github.com/Apolloplectic/hypoxic-burden-edf/releases)
    2. Unzip the folder
    3. Open terminal/command prompt in that folder
    4. Run: `pip install -r requirements.txt`
    
    ---
    
    #### **Step 3: Run the App**
    In terminal, run: `streamlit run app.py --server.maxUploadSize=4096`
    
    The `--server.maxUploadSize=4096` flag allows uploads up to **4 GB**.
    
    Your browser will open automatically with the app running locally.
    
    ---
    
    **Need help?** 
    - 📧 Email: `sam.johnson9797@gmail.com`
    - 🐙 GitHub Issues: [Report a problem](https://github.com/Apolloplectic/hypoxic-burden-edf/issues)
    """)

# --------------------------------------------------------------
# INITIALIZE SESSION STATE
# --------------------------------------------------------------
initialize_session_state()

# --------------------------------------------------------------
# DISPLAY ANALYSIS HISTORY IN SIDEBAR (Feature #6)
# --------------------------------------------------------------
display_analysis_history()

# --------------------------------------------------------------
# SETTING PRESETS
# --------------------------------------------------------------
}
PRESETS = {
    "Azarbarzin 2019 (Default)": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,      # Corrected: START not END
        'desat_end_sec': 90,       # Corrected: 90s not 120s
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'description': "Exact parameters from Azarbarzin et al. EHJ 2019"
    },
    "AASM 2023 Standard": {
        'pre_event_sec': 120,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 3,
        'artifact_filter': 'Mild (10%/s)',
        'preset_baseline': 0.0,
        'description': "Current clinical practice guidelines"
    },
    "Conservative (High Specificity)": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 4,
        'artifact_filter': 'Strict (5%/s)',
        'preset_baseline': 0.0,
        'description': "Minimizes false positives (4% threshold)"
    },
    "Aggressive (High Sensitivity)": {
        'pre_event_sec': 80,
        'desat_start_sec': 0,
        'desat_end_sec': 150,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'description': "Maximizes event detection"
    },
    "Custom": {
        'pre_event_sec': 100,
        'desat_start_sec': 0,
        'desat_end_sec': 90,
        'desat_threshold': 3,
        'artifact_filter': 'Off',
        'preset_baseline': 0.0,
        'description': "Manually configure all parameters"
    }
}

# =============================================
# SINGLE FILE ANALYSIS MODE
# =============================================
if edf_file is not None:
    # Load EDF file
    with st.spinner(f"📂 Loading {edf_file.name} ({edf_file.size / 1e6:.1f} MB)..."):
        raw, temp_path = load_edf_file(edf_file)
    
    if raw is None:
        st.error("❌ Failed to load EDF file. Please check the file format.")
        st.stop()
    
    st.success(f"✅ EDF loaded successfully! Duration: {raw.times[-1]/3600:.2f} hours")
    
    # Initialize analyzer
    analyzer = PSGAnalyzer(raw, temp_path)
    
    # --------------------------------------------------------------
    # DATA VALIDATION (Feature #3)
    # --------------------------------------------------------------
    st.markdown("#### 🔍 Data Quality Check")
    
    validator = PSGValidator(raw, analyzer)
    validation_results = validator.validate_all()
    
    # Display validation results
    validator.display_results(show_info=True)
    
    # Stop if critical errors
    if not validation_results['valid']:
        st.error("**Cannot proceed with analysis due to critical errors above.**")
        st.stop()
    
    # Display detected channels
    st.markdown("---")
    st.markdown("#### 📡 Detected Channels")
    st.write(f"**SpO₂:** `{analyzer.spo2_ch or 'Not found ❌'}`")
    st.write(f"**Airflow:** `{analyzer.flow_ch or 'Not found (will use SpO₂-based detection)'}`")
    st.write(f"**EEG:** `{analyzer.eeg_ch or 'Not found (staging limited)'}`")
    
    # --------------------------------------------------------------
    # MANUAL CHANNEL SELECTION (Feature #8)
    # --------------------------------------------------------------
    if not analyzer.spo2_ch or not analyzer.flow_ch or not analyzer.eeg_ch:
        with st.expander("🔧 Manual Channel Selection", expanded=not analyzer.spo2_ch):
            st.markdown("**Auto-detection failed for some channels. Select manually:**")
            
            all_channels = raw.ch_names
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if not analyzer.spo2_ch:
                    manual_spo2 = st.selectbox(
                        "SpO₂ Channel",
                        ["None"] + all_channels,
                        help="Required for analysis"
                    )
                    if manual_spo2 != "None":
                        analyzer.spo2_ch = manual_spo2
                        st.success(f"✅ Set to: {manual_spo2}")
            
            with col2:
                if not analyzer.flow_ch:
                    manual_flow = st.selectbox(
                        "Airflow Channel (optional)",
                        ["None"] + all_channels,
                        help="Improves event detection accuracy"
                    )
                    if manual_flow != "None":
                        analyzer.flow_ch = manual_flow
                        st.success(f"✅ Set to: {manual_flow}")
            
            with col3:
                if not analyzer.eeg_ch:
                    manual_eeg = st.selectbox(
                        "EEG Channel (optional)",
                        ["None"] + all_channels,
                        help="Enables sleep staging"
                    )
                    if manual_eeg != "None":
                        analyzer.eeg_ch = manual_eeg
                        st.success(f"✅ Set to: {manual_eeg}")
            
            st.info("💡 **Common alternate names:**\n"
                   "- SpO₂: 'Oxygen', 'O2', 'SAT', 'SpO2'\n"
                   "- Airflow: 'Flow', 'Nasal', 'Cannula', 'Pressure'\n"
                   "- EEG: 'C3', 'C4', 'F3', 'F4', 'EEG1', 'EEG2'")
    
    if not analyzer.spo2_ch:
        st.error("❌ SpO₂ channel is required for analysis.")
        st.stop()
    
    # Check for MIT annotations
    if analyzer.check_mit_annotations():
        use_mit = st.checkbox(
            "✨ Use MIT Gold Standard Annotations",
            value=True,
            help="MIT-annotated sleep stages and events from SHHS/slpdb database"
        )
        st.session_state.use_mit_st = use_mit
        if use_mit:
            st.success(f"🎯 MIT annotations loaded: {len(analyzer.manual_events)} events, AHI = {analyzer.manual_ahi:.1f}")
    else:
        st.info("ℹ️ No MIT annotations found — using automated detection")
        st.session_state.use_mit_st = False
    
    # --------------------------------------------------------------
    # PRESET COMPARISON TOOL (Feature #1)
    # --------------------------------------------------------------
    st.markdown("---")
    run_comparison = st.checkbox(
        "🔬 Compare Multiple Presets",
        value=False,
        help="Run analysis with multiple presets to see how methodology affects results"
    )
    
    if run_comparison:
        st.info("**Preset Comparison Mode:** Select presets to compare below. Analysis will run for each selected preset.")
        
        comparison_presets = st.multiselect(
            "Select presets to compare:",
            [p for p in PRESETS.keys() if p != "Custom"],
            default=["Azarbarzin 2019 (Default)", "AASM 2023 Standard"],
            help="Choose 2-4 presets to compare"
        )
        
        if len(comparison_presets) < 2:
            st.warning("⚠️ Please select at least 2 presets to compare")
        elif len(comparison_presets) > 4:
            st.warning("⚠️ Maximum 4 presets for comparison (performance)")
    
    # --------------------------------------------------------------
    # ADVANCED SETTINGS WITH PRESETS
    # --------------------------------------------------------------
    with st.expander("⚙️ Advanced Settings", expanded=False):
        # Preset selector
        preset_choice = st.selectbox(
            "📋 Analysis Preset",
            list(PRESETS.keys()),
            index=0,
            help="Choose a validated preset or customize all parameters"
        )
        
        # Show preset description
        st.info(f"ℹ️ **{preset_choice}:** {PRESETS[preset_choice]['description']}")
        
        # Load preset values
        params = PRESETS[preset_choice].copy()
        
        st.markdown("---")
        
        # If Custom, show sliders; otherwise show read-only values
        if preset_choice == "Custom":
            st.markdown("#### 🎚️ Event Detection Parameters")
            col1, col2 = st.columns(2)
            
            with col1:
                params['pre_event_sec'] = st.slider(
                    "Pre-event baseline window (s)",
                    min_value=30,
                    max_value=180,
                    value=params['pre_event_sec'],
                    step=1,
                    help="Time before event to calculate baseline SpO₂"
                )
                params['desat_start_sec'] = st.slider(
                    "Desaturation start (s before event end)",
                    min_value=15,
                    max_value=120,
                    value=params['desat_start_sec'],
                    step=1,
                    help="Start of desaturation window"
                )
            
            with col2:
                params['desat_end_sec'] = st.slider(
                    "Desaturation end (s after event end)",
                    min_value=60,
                    max_value=240,
                    value=params['desat_end_sec'],
                    step=1,
                    help="End of desaturation window (recovery time)"
                )
                params['artifact_filter'] = st.selectbox(
                    "SpO₂ artifact filter",
                    ["Off", "Mild (10%/s)", "Strict (5%/s)"],
                    index=["Off", "Mild (10%/s)", "Strict (5%/s)"].index(params['artifact_filter']),
                    help="Remove physiologically impossible SpO₂ changes"
                )
            
            st.markdown("#### 📊 Scoring Parameters")
            col3, col4 = st.columns(2)
            
            with col3:
                scoring_rule = st.selectbox(
                    "Desaturation threshold",
                    ["3% (AASM)", "4% (Legacy)"],
                    index=0 if params['desat_threshold'] == 3 else 1,
                    help="AASM recommends 3%"
                )
                params['desat_threshold'] = 3 if "3%" in scoring_rule else 4
            
            with col4:
                params['use_global_hb'] = st.checkbox(
                    "Calculate Global HB",
                    value=params['use_global_hb'],
                    help="Total oxygen debt over entire study"
                )
            
            if params['use_global_hb']:
                baseline_method = st.radio(
                    "Baseline SpO₂ method",
                    ["Automatic (95th percentile)", "Manual entry"],
                    help="Auto removes outliers/desaturations"
                )
                
                if baseline_method == "Manual entry":
                    params['preset_baseline'] = st.slider(
                        "Baseline SpO₂ (%)",
                        min_value=80.0,
                        max_value=100.0,
                        value=95.0,
                        step=0.1,
                        format="%.1f"
                    )
                else:
                    params['preset_baseline'] = 0.0
            else:
                params['preset_baseline'] = 0.0
        
        else:
            # Show preset values (read-only)
            st.markdown("#### 📋 Preset Configuration")
            
            st.markdown("##### Event-Specific HB Parameters:")
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**Pre-event baseline:** {params['pre_event_sec']}s before START")
                st.write(f"**Baseline method:** MAXIMUM SpO₂ (per Azarbarzin)")
                st.write(f"**Desat start offset:** {params['desat_start_sec']}s from START")
            
            with col2:
                st.write(f"**Desat end:** {params['desat_end_sec']}s after end")
                st.write(f"**Artifact filter:** {params['artifact_filter']}")
                st.write(f"**Desat threshold:** {params['desat_threshold']}%")
            
            st.markdown("##### Global HB Parameters:")
            baseline_text = 'Auto (95th percentile)' if params['preset_baseline'] == 0 else f'{params["preset_baseline"]:.1f}%'
            st.write(f"**Global baseline:** {baseline_text} (for whole-study burden)")
            st.caption("ℹ️ Note: Event-specific HB uses MAXIMUM in 100s before each event. Global HB uses this baseline for entire study.")
        
        
        # Store in session state
        st.session_state.analysis_params = params
        
        # Warnings for non-default settings
        if preset_choice == "Custom":
            if params['pre_event_sec'] != 100:
                st.warning(f"⚠️ Pre-event window: {params['pre_event_sec']}s (Azarbarzin default: 100s)")
            if params['desat_start_sec'] != 0 or params['desat_end_sec'] != 90:
                st.warning(f"⚠️ Desat window: {params['desat_start_sec']}s from START, {params['desat_end_sec']}s after END (Azarbarzin: 0s, 90s)")
            if params['desat_threshold'] != 3:
                st.warning("⚠️ Using 4% threshold (non-AASM)")
            if params['artifact_filter'] != "Off":
                st.warning(f"⚠️ Artifact filtering enabled: {params['artifact_filter']}")
        elif preset_choice != "Azarbarzin 2019 (Default)":
            st.info(f"✨ Using **{preset_choice}** - results may differ from Azarbarzin 2019 baseline")
    
    # Extract params for use
    pre_event_sec = params['pre_event_sec']
    desat_start_sec = params['desat_start_sec']
    desat_end_sec = params['desat_end_sec']
    artifact_filter = params['artifact_filter']
    desat_threshold = params['desat_threshold']
    preset_baseline = params['preset_baseline']
    # Note: Global HB is always calculated (use_global_hb always True)
    if desat_threshold == 4:
        st.warning("⚠️ Using 4% desaturation threshold (non-AASM standard)")
    if artifact_filter != "Off":
        st.info(f"ℹ️ Artifact filter enabled: {artifact_filter}")
    if not analyzer.flow_ch:
        st.warning("⚠️ No airflow channel found — AHI will be estimated from SpO₂ desaturations only")
    
    # --------------------------------------------------------------
    # ANALYSIS BUTTON
    # --------------------------------------------------------------
    st.markdown("---")
    
    if not st.session_state.analyzed:
        analyze_label = "🔬 Compare Presets" if run_comparison else "🚀 Analyze File"
        if st.button(analyze_label, type="primary", use_container_width=True):
            st.session_state.analyzed = True
            st.session_state.run_comparison = run_comparison
            if run_comparison:
                st.session_state.comparison_presets = comparison_presets
            st.rerun()
    else:
        # --------------------------------------------------------------
        # PRESET COMPARISON MODE (Feature #1)
        # --------------------------------------------------------------
        if st.session_state.get('run_comparison') and len(st.session_state.get('comparison_presets', [])) >= 2:
            st.markdown("---")
            st.subheader("🔬 Preset Comparison Results")
            
            comparison_results = {}
            
            with st.spinner("Running analysis with multiple presets..."):
                progress_bar = st.progress(0)
                
                for idx, preset_name in enumerate(st.session_state.comparison_presets):
                    status = st.empty()
                    status.text(f"Analyzing with {preset_name}... ({idx+1}/{len(st.session_state.comparison_presets)})")
                    
                    preset_params = PRESETS[preset_name]
                    
                    results = analyzer.run_full_analysis(
                        pre_event_sec=preset_params['pre_event_sec'],
                        desat_start_sec=preset_params['desat_start_sec'],
                        desat_end_sec=preset_params['desat_end_sec'],
                        artifact_filter=preset_params['artifact_filter'],
                        desat_threshold=preset_params['desat_threshold'],
                        use_global_hb=True,
                        preset_baseline=preset_params['preset_baseline'],
                        use_mit_st=st.session_state.use_mit_st
                    )
                    
                    comparison_results[preset_name] = results
                    progress_bar.progress((idx + 1) / len(st.session_state.comparison_presets))
                
                status.empty()
                progress_bar.empty()
            
            # Display comparison table
            st.markdown("#### 📊 Comparison Table")
            
            comparison_data = []
            for preset_name, results in comparison_results.items():
                comparison_data.append({
                    'Preset': preset_name,
                    'AHI': f"{results['ahi']:.1f}",
                    'ODI': f"{results['odi']:.1f}",
                    'Obstructive HB': f"{results['total_hb']:.1f}",
                    'CI Lower': f"{results['ci'][0]:.1f}",
                    'CI Upper': f"{results['ci'][1]:.1f}",
                    'Risk': 'Low' if results['total_hb'] < 20 else 'Moderate' if results['total_hb'] < 53 else 'High' if results['total_hb'] < 88 else 'Very High'
                })
            
            df_comparison = pd.DataFrame(comparison_data)
            st.dataframe(df_comparison, use_container_width=True, hide_index=True)
            
            # Visualization
            st.markdown("#### 📈 Visual Comparison")
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
            
            presets = [d['Preset'] for d in comparison_data]
            ahis = [float(d['AHI']) for d in comparison_data]
            hbs = [float(d['Obstructive HB']) for d in comparison_data]
            
            # AHI comparison
            colors = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000']
            ax1.bar(range(len(presets)), ahis, color=colors[:len(presets)])
            ax1.set_xticks(range(len(presets)))
            ax1.set_xticklabels([p.split('(')[0].strip() for p in presets], rotation=15, ha='right')
            ax1.set_ylabel('AHI (events/hour)')
            ax1.set_title('AHI by Preset')
            ax1.grid(axis='y', alpha=0.3)
            
            # HB comparison
            ax2.bar(range(len(presets)), hbs, color=colors[:len(presets)])
            ax2.set_xticks(range(len(presets)))
            ax2.set_xticklabels([p.split('(')[0].strip() for p in presets], rotation=15, ha='right')
            ax2.set_ylabel('Hypoxic Burden (%min/h)')
            ax2.set_title('Obstructive HB by Preset')
            ax2.grid(axis='y', alpha=0.3)
            
            plt.tight_layout()
            st.pyplot(fig)
            plt.close(fig)
            
            # Key insights
            st.markdown("#### 💡 Key Insights")
            
            ahi_range = max(ahis) - min(ahis)
            hb_range = max(hbs) - min(hbs)
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("AHI Range", f"{ahi_range:.1f}", help="Difference between highest and lowest AHI")
            with col2:
                st.metric("HB Range", f"{hb_range:.1f}", help="Difference between highest and lowest HB")
            
            if ahi_range > 5:
                st.warning(f"⚠️ **Wide AHI variation ({ahi_range:.1f} events/hr)** - Methodology significantly affects event detection. "
                          "Conservative preset may undercount, Aggressive may overcount.")
            else:
                st.success(f"✅ **Consistent AHI across presets ({ahi_range:.1f} range)** - Results are robust to methodology changes.")
            
            # --------------------------------------------------------------
            # EXCEL EXPORT FOR COMPARISON (Feature #9 Enhancement)
            # --------------------------------------------------------------
            st.markdown("---")
            st.markdown("#### 📊 Export Comparison")
            
            if st.button("📥 Download Comparison as Excel", use_container_width=True):
                with st.spinner("Generating comparison Excel workbook..."):
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    wb = Workbook()
                    
                    # Header styling
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    # Sheet 1: Comparison Summary
                    ws_summary = wb.active
                    ws_summary.title = "Comparison Summary"
                    
                    ws_summary['A1'] = "Preset Comparison Analysis"
                    ws_summary['A1'].font = Font(bold=True, size=14)
                    ws_summary['A2'] = f"File: {edf_file.name}"
                    ws_summary['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    
                    # Comparison table
                    ws_summary['A5'] = "Preset"
                    ws_summary['B5'] = "AHI"
                    ws_summary['C5'] = "ODI"
                    ws_summary['D5'] = "Obstructive HB"
                    ws_summary['E5'] = "CI Lower"
                    ws_summary['F5'] = "CI Upper"
                    ws_summary['G5'] = "Risk Level"
                    
                    for col in ['A5', 'B5', 'C5', 'D5', 'E5', 'F5', 'G5']:
                        ws_summary[col].fill = header_fill
                        ws_summary[col].font = header_font
                    
                    for idx, data in enumerate(comparison_data, 6):
                        ws_summary[f'A{idx}'] = data['Preset']
                        ws_summary[f'B{idx}'] = float(data['AHI'])
                        ws_summary[f'C{idx}'] = float(data['ODI'])
                        ws_summary[f'D{idx}'] = float(data['Obstructive HB'])
                        ws_summary[f'E{idx}'] = float(data['CI Lower'])
                        ws_summary[f'F{idx}'] = float(data['CI Upper'])
                        ws_summary[f'G{idx}'] = data['Risk']
                    
                    # Auto-adjust columns
                    ws_summary.column_dimensions['A'].width = 30
                    for col in ['B', 'C', 'D', 'E', 'F']:
                        ws_summary.column_dimensions[col].width = 15
                    ws_summary.column_dimensions['G'].width = 15
                    
                    # Create a sheet for each preset's detailed results
                    for preset_name, preset_results in comparison_results.items():
                        # Shorten sheet name (Excel limit is 31 chars)
                        sheet_name = preset_name.replace(" (Default)", "").replace(" (High Specificity)", "").replace(" (High Sensitivity)", "")[:31]
                        ws = wb.create_sheet(sheet_name)
                        
                        ws['A1'] = f"{preset_name} - Detailed Results"
                        ws['A1'].font = Font(bold=True, size=12)
                        
                        # Main metrics
                        ws['A3'] = "Metric"
                        ws['B3'] = "Value"
                        ws['A3'].fill = header_fill
                        ws['B3'].fill = header_fill
                        ws['A3'].font = header_font
                        ws['B3'].font = header_font
                        
                        ws['A4'] = "Duration (h)"
                        ws['B4'] = preset_results['duration']
                        ws['A5'] = "AHI"
                        ws['B5'] = preset_results['ahi']
                        ws['A6'] = "ODI"
                        ws['B6'] = preset_results['odi']
                        ws['A7'] = "Obstructive HB"
                        ws['B7'] = preset_results['total_hb']
                        ws['A8'] = "95% CI Lower"
                        ws['B8'] = preset_results['ci'][0]
                        ws['A9'] = "95% CI Upper"
                        ws['B9'] = preset_results['ci'][1]
                        
                        if preset_results.get('global_hb'):
                            ws['A10'] = "Global HB"
                            ws['B10'] = preset_results['global_hb']
                            ws['A11'] = "Baseline SpO₂"
                            ws['B11'] = preset_results['baseline_used']
                        
                        # Stage-specific results if available
                        if preset_results.get('stage_hb'):
                            ws['A13'] = "Stage-Specific Results"
                            ws['A13'].font = Font(bold=True, size=11)
                            
                            ws['A14'] = "Stage"
                            ws['B14'] = "Time (h)"
                            ws['C14'] = "AHI"
                            ws['D14'] = "ODI"
                            ws['E14'] = "HB"
                            
                            for col in ['A14', 'B14', 'C14', 'D14', 'E14']:
                                ws[col].fill = header_fill
                                ws[col].font = header_font
                            
                            row = 15
                            for stage in ['W', 'N1', 'N2', 'N3', 'REM']:
                                if stage in preset_results['stage_hb']:
                                    data = preset_results['stage_hb'][stage]
                                    ws[f'A{row}'] = stage
                                    ws[f'B{row}'] = data['hrs']
                                    ws[f'C{row}'] = data['AHI']
                                    ws[f'D{row}'] = data['ODI']
                                    ws[f'E{row}'] = data['HB']
                                    row += 1
                        
                        # Auto-adjust columns
                        ws.column_dimensions['A'].width = 20
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 12
                        ws.column_dimensions['D'].width = 12
                        ws.column_dimensions['E'].width = 12
                    
                    # Parameters sheet
                    ws_params = wb.create_sheet("Parameters Used")
                    ws_params['A1'] = "Analysis Parameters by Preset"
                    ws_params['A1'].font = Font(bold=True, size=12)
                    
                    ws_params['A3'] = "Parameter"
                    for idx, preset_name in enumerate(comparison_results.keys(), 2):
                        col_letter = chr(65 + idx)  # B, C, D, etc.
                        ws_params[f'{col_letter}3'] = preset_name.split('(')[0].strip()
                        ws_params[f'{col_letter}3'].fill = header_fill
                        ws_params[f'{col_letter}3'].font = header_font
                    
                    ws_params['A3'].fill = header_fill
                    ws_params['A3'].font = header_font
                    
                    param_names = [
                        ("Pre-event baseline (s)", 'pre_event_sec'),
                        ("Desat start (s)", 'desat_start_sec'),
                        ("Desat end (s)", 'desat_end_sec'),
                        ("Desat threshold (%)", 'desat_threshold'),
                        ("Artifact filter", 'artifact_filter'),
                        ("Global HB enabled", 'use_global_hb')
                    ]
                    
                    for row_idx, (param_name, param_key) in enumerate(param_names, 4):
                        ws_params[f'A{row_idx}'] = param_name
                        for col_idx, preset_name in enumerate(comparison_results.keys(), 2):
                            col_letter = chr(65 + col_idx)
                            preset_params = PRESETS[preset_name]
                            value = preset_params[param_key]
                            if param_key == 'use_global_hb':
                                value = "Yes" if value else "No"
                            ws_params[f'{col_letter}{row_idx}'] = value
                    
                    ws_params.column_dimensions['A'].width = 25
                    for col_idx in range(len(comparison_results)):
                        col_letter = chr(66 + col_idx)
                        ws_params.column_dimensions[col_letter].width = 20
                    
                    # Save to buffer
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Download Comparison Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"HB_Comparison_{edf_file.name.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
            
            # Use first preset results for detailed display below
            results = list(comparison_results.values())[0]
            params = PRESETS[list(comparison_results.keys())[0]]
            
            st.markdown("---")
            st.markdown("### 📊 Detailed Results (First Preset)")
        else:
            # Regular single analysis
            # Run analysis
            with st.spinner("🔬 Analyzing PSG data..."):
                results = analyzer.run_full_analysis(
                    pre_event_sec=pre_event_sec,
                    desat_start_sec=desat_start_sec,
                    desat_end_sec=desat_end_sec,
                    artifact_filter=artifact_filter,
                    desat_threshold=desat_threshold,
                    use_global_hb=True,  # Always calculate Global HB
                    preset_baseline=preset_baseline,
                    use_mit_st=st.session_state.use_mit_st
                )
        
        # --------------------------------------------------------------
        # SAVE TO SESSION HISTORY (Feature #6)
        # --------------------------------------------------------------
        save_analysis_to_session(edf_file.name, results, params)
        
        # Display results
        st.markdown("---")
        st.subheader("📊 Analysis Results")
        
        # --------------------------------------------------------------
        # COMPARISON WITH PREVIOUS (Feature #6)
        # --------------------------------------------------------------
        previous_analysis = compare_with_previous()
        if previous_analysis:
            display_comparison(results, previous_analysis)
            st.markdown("---")
        
        # Main metrics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                "AHI",
                f"{results['ahi']:.1f}",
                help="Apnea-Hypopnea Index (events per hour)"
            )
            if results.get('manual_ahi') is not None:
                delta = results['ahi'] - results['manual_ahi']
                st.caption(f"MIT Gold Std: {results['manual_ahi']:.1f} (Δ {delta:+.1f})")
        
        with col2:
            st.metric(
                f"ODI ({desat_threshold}%)",
                f"{results['odi']:.1f}",
                help=f"Oxygen Desaturation Index (≥{desat_threshold}% drops per hour)"
            )
        
        with col3:
            if len(results['events']) > 0:
                ci_str = f"[{results['ci'][0]:.1f}–{results['ci'][1]:.1f}]"
                st.metric(
                    "Obstructive HB",
                    f"{results['total_hb']:.1f}",
                    help=f"Event-specific Hypoxic Burden. 95% CI: {ci_str}"
                )
                st.caption(f"95% CI: {ci_str}")
            else:
                st.metric("Obstructive HB", "0.0")
        
        # Risk level
        risk_level = "Low"
        if results['total_hb'] >= 88:
            risk_level = "Very High"
            risk_color = "🔴"
        elif results['total_hb'] >= 53:
            risk_level = "High"
            risk_color = "🟠"
        elif results['total_hb'] >= 20:
            risk_level = "Moderate"
            risk_color = "🟡"
        else:
            risk_color = "🟢"
        
        st.markdown(f"### {risk_color} Risk Level: **{risk_level}**")
        
        # Global HB (if calculated)
        if results.get('global_hb') is not None:
            st.markdown("---")
            st.subheader("🌍 Global Hypoxic Burden")
            col1, col2 = st.columns(2)
            with col1:
                st.metric(
                    "Global HB",
                    f"{results['global_hb']:.2f} (%min)/h",
                    help="Total oxygen debt over entire sleep study"
                )
            with col2:
                st.metric(
                    "Baseline SpO₂",
                    f"{results['baseline_used']:.1f}%",
                    help="SpO₂ baseline used for calculation"
                )
        
        # Stage-specific results
        if results['stage_hb']:
            st.markdown("---")
            st.subheader("😴 Stage-Specific Metrics")
            
            stage_data = []
            for stage in ['W', 'N1', 'N2', 'N3', 'REM']:
                if stage in results['stage_hb']:
                    data = results['stage_hb'][stage]
                    stage_data.append({
                        'Stage': stage,
                        'Time (h)': f"{data['hrs']:.1f}",
                        'AHI': f"{data['AHI']:.1f}",
                        'ODI': f"{data['ODI']:.1f}",
                        'HB': f"{data['HB']:.2f}"
                    })
            
            if stage_data:
                import pandas as pd
                st.dataframe(pd.DataFrame(stage_data), use_container_width=True)
        
        # Report generation
        st.markdown("---")
        st.subheader("📄 Generate Report")
        
        col1, col2 = st.columns(2)
        with col1:
            proof_mode = st.selectbox(
                "Include proof plots",
                ["None", "Overlay (Azarbarzin-style)", "Full (all events)"],
                index=1
            )
        with col2:
            include_stages = st.checkbox("Include stage-specific results", value=True)
        
        if PDF_AVAILABLE and st.button("📥 Download PDF Report", type="primary", use_container_width=True):
            with st.spinner("Generating PDF report..."):
                pdf_generator = PDFReportGenerator()
                buffer = pdf_generator.generate_report(
                    filename=edf_file.name,
                    results=results,
                    proof_mode=proof_mode,
                    include_stages=include_stages
                )
                
                st.download_button(
                    label="⬇️ Download Report",
                    data=buffer.getvalue(),
                    file_name=f"HB_Report_{edf_file.name.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    type="primary",
                    use_container_width=True
                )
        elif not PDF_AVAILABLE:
            st.info("ℹ️ PDF export currently unavailable. Use Excel export below instead.")
        
        # --------------------------------------------------------------
        # EXCEL EXPORT (Feature #9)
        # --------------------------------------------------------------
        st.markdown("---")
        st.markdown("#### 📊 Excel Export (Multi-Sheet)")
        
        if st.button("📥 Generate Excel Report", use_container_width=True):
            with st.spinner("Generating Excel workbook..."):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                
                # Sheet 1: Summary
                ws_summary = wb.active
                ws_summary.title = "Summary"
                
                # Header styling
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # Summary data
                ws_summary['A1'] = "Hypoxic Burden Analysis Report"
                ws_summary['A1'].font = Font(bold=True, size=14)
                ws_summary['A3'] = "File"
                ws_summary['B3'] = edf_file.name
                ws_summary['A4'] = "Analysis Date"
                ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ws_summary['A5'] = "Duration (hours)"
                ws_summary['B5'] = results['duration']
                
                ws_summary['A7'] = "Metric"
                ws_summary['B7'] = "Value"
                ws_summary['C7'] = "95% CI"
                ws_summary['A7'].fill = header_fill
                ws_summary['B7'].fill = header_fill
                ws_summary['C7'].fill = header_fill
                ws_summary['A7'].font = header_font
                ws_summary['B7'].font = header_font
                ws_summary['C7'].font = header_font
                
                ws_summary['A8'] = "AHI"
                ws_summary['B8'] = results['ahi']
                ws_summary['A9'] = "ODI"
                ws_summary['B9'] = results['odi']
                ws_summary['A10'] = "Obstructive HB"
                ws_summary['B10'] = results['total_hb']
                ws_summary['C10'] = f"[{results['ci'][0]:.1f} - {results['ci'][1]:.1f}]"
                
                if results.get('global_hb'):
                    ws_summary['A11'] = "Global HB"
                    ws_summary['B11'] = results['global_hb']
                    ws_summary['A12'] = "Baseline SpO₂"
                    ws_summary['B12'] = results['baseline_used']
                
                # Auto-adjust column widths
                ws_summary.column_dimensions['A'].width = 25
                ws_summary.column_dimensions['B'].width = 15
                ws_summary.column_dimensions['C'].width = 20
                
                # Sheet 2: Stage-Specific Results
                if results['stage_hb']:
                    ws_stages = wb.create_sheet("Stage-Specific")
                    
                    stage_data = []
                    for stage in ['W', 'N1', 'N2', 'N3', 'REM']:
                        if stage in results['stage_hb']:
                            data = results['stage_hb'][stage]
                            stage_data.append({
                                'Stage': stage,
                                'Time (h)': data['hrs'],
                                'AHI': data['AHI'],
                                'ODI': data['ODI'],
                                'HB': data['HB']
                            })
                    
                    df_stages = pd.DataFrame(stage_data)
                    
                    # Add header
                    for r_idx, row in enumerate(dataframe_to_rows(df_stages, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws_stages.cell(row=r_idx, column=c_idx, value=value)
                            if r_idx == 1:  # Header row
                                cell.fill = header_fill
                                cell.font = header_font
                    
                    # Auto-adjust columns
                    for column in ws_stages.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws_stages.column_dimensions[column_letter].width = max_length + 2
                
                # Sheet 3: Analysis Parameters
                ws_params = wb.create_sheet("Parameters")
                ws_params['A1'] = "Analysis Parameters"
                ws_params['A1'].font = Font(bold=True, size=14)
                
                ws_params['A3'] = "Parameter"
                ws_params['B3'] = "Value"
                ws_params['A3'].fill = header_fill
                ws_params['B3'].fill = header_fill
                ws_params['A3'].font = header_font
                ws_params['B3'].font = header_font
                
                params_list = [
                    ("Pre-event baseline (s)", params['pre_event_sec']),
                    ("Desat start (s)", params['desat_start_sec']),
                    ("Desat end (s)", params['desat_end_sec']),
                    ("Desat threshold (%)", params['desat_threshold']),
                    ("Artifact filter", params['artifact_filter']),
                    ("Global HB enabled", "Yes" if params['use_global_hb'] else "No")
                ]
                
                for idx, (param, value) in enumerate(params_list, 4):
                    ws_params[f'A{idx}'] = param
                    ws_params[f'B{idx}'] = value
                
                ws_params.column_dimensions['A'].width = 25
                ws_params.column_dimensions['B'].width = 15
                
                # Save to buffer
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=excel_buffer.getvalue(),
                    file_name=f"HB_Analysis_{edf_file.name.replace('.edf', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
        
        # Reset button
        if st.button("🔄 Analyze Another File", use_container_width=True):
            st.session_state.analyzed = False
            if os.path.exists(temp_path):
                os.remove(temp_path)
            st.rerun()

# =============================================
# TREATMENT COMPARISON MODE
# =============================================
elif analysis_mode == "Treatment Comparison (Before/After)" and pre_treatment_file is not None and post_treatment_file is not None:
    st.markdown("---")
    st.markdown("## 🔬 Treatment Comparison Analysis")
    
    # Load both files
    with st.spinner("📂 Loading pre-treatment PSG..."):
        raw_pre, temp_path_pre = load_edf_file(pre_treatment_file)
    
    with st.spinner("📂 Loading post-treatment PSG..."):
        raw_post, temp_path_post = load_edf_file(post_treatment_file)
    
    if raw_pre is None or raw_post is None:
        st.error("❌ Failed to load one or both EDF files. Please check the file formats.")
        st.stop()
    
    # Initialize analyzers
    analyzer_pre = PSGAnalyzer(raw_pre, temp_path_pre)
    analyzer_post = PSGAnalyzer(raw_post, temp_path_post)
    
    # Validate both files
    st.markdown("### 🔍 Data Quality Check")
    
    # Just show basic info, don't validate
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### Pre-Treatment")
        st.write(f"✅ Loaded: {pre_treatment_file.name}")
        st.write(f"Duration: {raw_pre.times[-1]/3600:.2f} hours")
        st.write(f"SpO₂: {analyzer_pre.spo2_ch or 'Not detected'}")
    
    with col2:
        st.markdown("#### Post-Treatment")
        st.write(f"✅ Loaded: {post_treatment_file.name}")
        st.write(f"Duration: {raw_post.times[-1]/3600:.2f} hours")
        st.write(f"SpO₂: {analyzer_post.spo2_ch or 'Not detected'}")
    
    # Check for critical issues
    if not analyzer_pre.spo2_ch or not analyzer_post.spo2_ch:
        st.error("❌ SpO₂ channel not found in one or both files. Cannot proceed.")
        st.stop()
    
    # Select preset for analysis
    st.markdown("---")
    st.markdown("### ⚙️ Analysis Settings")
    
    comparison_preset = st.selectbox(
        "Analysis Preset (applied to both files)",
        list(PRESETS.keys()),
        index=0,
        help="Use same preset for fair comparison"
    )
    
    params = PRESETS[comparison_preset].copy()
    st.info(f"ℹ️ **{comparison_preset}:** {params['description']}")
    
    # Run analysis button
    if st.button("🚀 Run Comparison Analysis", type="primary", use_container_width=True):
        
        # Analyze pre-treatment
        with st.spinner("🔬 Analyzing pre-treatment PSG..."):
            results_pre = analyzer_pre.run_full_analysis(
                pre_event_sec=params['pre_event_sec'],
                desat_start_sec=params['desat_start_sec'],
                desat_end_sec=params['desat_end_sec'],
                artifact_filter=params['artifact_filter'],
                desat_threshold=params['desat_threshold'],
                use_global_hb=True,
                preset_baseline=params['preset_baseline'],
                use_mit_st=st.session_state.use_mit_st
            )
        
        # Analyze post-treatment
        with st.spinner("🔬 Analyzing post-treatment PSG..."):
            results_post = analyzer_post.run_full_analysis(
                pre_event_sec=params['pre_event_sec'],
                desat_start_sec=params['desat_start_sec'],
                desat_end_sec=params['desat_end_sec'],
                artifact_filter=params['artifact_filter'],
                desat_threshold=params['desat_threshold'],
                use_global_hb=True,
                preset_baseline=params['preset_baseline'],
                use_mit_st=st.session_state.use_mit_st
            )
        
        # Display comparison results
        st.markdown("---")
        st.markdown("## 📊 Treatment Effect Summary")
        
        # Overall metrics comparison
        st.markdown("### Primary Outcomes")
        
        comparison_data = {
            'Metric': ['AHI (events/h)', 'ODI (events/h)', 'Hypoxic Burden (%·min/h)', 'Global HB (%·min/h)'],
            'Pre-Treatment': [
                f"{results_pre['ahi']:.1f}",
                f"{results_pre['odi']:.1f}",
                f"{results_pre['total_hb']:.1f}",
                f"{results_pre.get('global_hb', 0):.1f}"
            ],
            'Post-Treatment': [
                f"{results_post['ahi']:.1f}",
                f"{results_post['odi']:.1f}",
                f"{results_post['total_hb']:.1f}",
                f"{results_post.get('global_hb', 0):.1f}"
            ],
            'Absolute Change': [
                f"{results_post['ahi'] - results_pre['ahi']:.1f}",
                f"{results_post['odi'] - results_pre['odi']:.1f}",
                f"{results_post['total_hb'] - results_pre['total_hb']:.1f}",
                f"{results_post.get('global_hb', 0) - results_pre.get('global_hb', 0):.1f}"
            ],
            '% Improvement': []
        }
        
        # Calculate % improvement
        for i, pre_val in enumerate([results_pre['ahi'], results_pre['odi'], results_pre['total_hb'], results_pre.get('global_hb', 0)]):
            post_val = [results_post['ahi'], results_post['odi'], results_post['total_hb'], results_post.get('global_hb', 0)][i]
            if pre_val > 0:
                improvement = ((pre_val - post_val) / pre_val) * 100
                comparison_data['% Improvement'].append(f"{improvement:.1f}%")
            else:
                comparison_data['% Improvement'].append("N/A")
        
        df_comparison = pd.DataFrame(comparison_data)
        st.dataframe(df_comparison, use_container_width=True, hide_index=True)
        
        # Treatment efficacy interpretation
        ahi_improvement = ((results_pre['ahi'] - results_post['ahi']) / results_pre['ahi'] * 100) if results_pre['ahi'] > 0 else 0
        hb_improvement = ((results_pre['total_hb'] - results_post['total_hb']) / results_pre['total_hb'] * 100) if results_pre['total_hb'] > 0 else 0
        
        st.markdown("### 🎯 Treatment Efficacy")
        
        if ahi_improvement >= 50 and hb_improvement >= 50:
            st.success(f"✅ **Excellent Response**: AHI improved {ahi_improvement:.1f}%, HB improved {hb_improvement:.1f}%")
        elif ahi_improvement >= 30 and hb_improvement >= 30:
            st.info(f"✓ **Good Response**: AHI improved {ahi_improvement:.1f}%, HB improved {hb_improvement:.1f}%")
        elif ahi_improvement >= 10 and hb_improvement >= 10:
            st.warning(f"⚠️ **Partial Response**: AHI improved {ahi_improvement:.1f}%, HB improved {hb_improvement:.1f}%")
        else:
            st.error(f"❌ **Poor Response**: AHI improved {ahi_improvement:.1f}%, HB improved {hb_improvement:.1f}%")
        
        # Stage-by-stage comparison
        st.markdown("---")
        st.markdown("### 📊 Stage-Specific Comparison")
        
        # Build stage comparison table
        stage_comparison_rows = []
        for stage in ['REM', 'N3', 'N2', 'N1', 'W']:
            if stage in results_pre['stage_hb'] and stage in results_post['stage_hb']:
                pre_stage = results_pre['stage_hb'][stage]
                post_stage = results_post['stage_hb'][stage]
                
                # Calculate improvements
                ahi_change = post_stage['AHI'] - pre_stage['AHI']
                odi_change = post_stage['ODI'] - pre_stage['ODI']
                hb_change = post_stage['HB'] - pre_stage['HB']
                
                hb_improvement_pct = ((pre_stage['HB'] - post_stage['HB']) / pre_stage['HB'] * 100) if pre_stage['HB'] > 0 else 0
                
                stage_comparison_rows.append({
                    'Stage': stage,
                    'Pre-AHI': f"{pre_stage['AHI']:.1f}",
                    'Post-AHI': f"{post_stage['AHI']:.1f}",
                    'AHI Δ': f"{ahi_change:.1f}",
                    'Pre-HB': f"{pre_stage['HB']:.1f}",
                    'Post-HB': f"{post_stage['HB']:.1f}",
                    'HB Improvement': f"{hb_improvement_pct:.1f}%"
                })
        
        if stage_comparison_rows:
            df_stage_comparison = pd.DataFrame(stage_comparison_rows)
            st.dataframe(df_stage_comparison, use_container_width=True, hide_index=True)
        
        # Visual comparison charts
        st.markdown("---")
        st.markdown("### 📈 Visual Comparison")
        
        # Create comparison bar chart
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # AHI/ODI comparison
        metrics = ['AHI', 'ODI']
        pre_values = [results_pre['ahi'], results_pre['odi']]
        post_values = [results_post['ahi'], results_post['odi']]
        
        x = np.arange(len(metrics))
        width = 0.35
        
        ax1.bar(x - width/2, pre_values, width, label='Pre-Treatment', color='#e74c3c', alpha=0.8)
        ax1.bar(x + width/2, post_values, width, label='Post-Treatment', color='#27ae60', alpha=0.8)
        ax1.set_ylabel('Events per hour')
        ax1.set_title('AHI & ODI Comparison')
        ax1.set_xticks(x)
        ax1.set_xticklabels(metrics)
        ax1.legend()
        ax1.grid(axis='y', alpha=0.3)
        
        # HB comparison
        hb_metrics = ['Event-Specific HB', 'Global HB']
        hb_pre_values = [results_pre['total_hb'], results_pre.get('global_hb', 0)]
        hb_post_values = [results_post['total_hb'], results_post.get('global_hb', 0)]
        
        x2 = np.arange(len(hb_metrics))
        
        ax2.bar(x2 - width/2, hb_pre_values, width, label='Pre-Treatment', color='#e74c3c', alpha=0.8)
        ax2.bar(x2 + width/2, hb_post_values, width, label='Post-Treatment', color='#27ae60', alpha=0.8)
        ax2.set_ylabel('(%·min)/h')
        ax2.set_title('Hypoxic Burden Comparison')
        ax2.set_xticks(x2)
        ax2.set_xticklabels(hb_metrics, rotation=15, ha='right')
        ax2.legend()
        ax2.grid(axis='y', alpha=0.3)
        
        plt.tight_layout()
        st.pyplot(fig)
        
        # Download comparison report
        st.markdown("---")
        st.markdown("### 📥 Export Comparison Report")
        
        # Create comprehensive Excel workbook
        from io import BytesIO
        
        # Prepare data for Excel export
        excel_buffer = BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Sheet 1: Summary Comparison
            summary_df = pd.DataFrame({
                'Metric': ['AHI (events/h)', 'ODI (events/h)', 'Event-Specific HB (%·min/h)', 
                          'Global HB (%·min/h)', 'Duration (hours)', 'Total Events'],
                'Pre-Treatment': [
                    results_pre['ahi'],
                    results_pre['odi'],
                    results_pre['total_hb'],
                    results_pre.get('global_hb', 0),
                    results_pre['duration'],
                    len(results_pre.get('events', []))
                ],
                'Post-Treatment': [
                    results_post['ahi'],
                    results_post['odi'],
                    results_post['total_hb'],
                    results_post.get('global_hb', 0),
                    results_post['duration'],
                    len(results_post.get('events', []))
                ],
                'Absolute Change': [
                    results_post['ahi'] - results_pre['ahi'],
                    results_post['odi'] - results_pre['odi'],
                    results_post['total_hb'] - results_pre['total_hb'],
                    results_post.get('global_hb', 0) - results_pre.get('global_hb', 0),
                    results_post['duration'] - results_pre['duration'],
                    len(results_post.get('events', [])) - len(results_pre.get('events', []))
                ],
                '% Change': [
                    ((results_post['ahi'] - results_pre['ahi']) / results_pre['ahi'] * 100) if results_pre['ahi'] > 0 else 0,
                    ((results_post['odi'] - results_pre['odi']) / results_pre['odi'] * 100) if results_pre['odi'] > 0 else 0,
                    ((results_post['total_hb'] - results_pre['total_hb']) / results_pre['total_hb'] * 100) if results_pre['total_hb'] > 0 else 0,
                    ((results_post.get('global_hb', 0) - results_pre.get('global_hb', 0)) / results_pre.get('global_hb', 1) * 100) if results_pre.get('global_hb', 0) > 0 else 0,
                    0,
                    0
                ]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 2: Stage-by-Stage Comparison
            stage_rows = []
            for stage in ['REM', 'N3', 'N2', 'N1', 'W']:
                if stage in results_pre.get('stage_hb', {}) and stage in results_post.get('stage_hb', {}):
                    pre_stage = results_pre['stage_hb'][stage]
                    post_stage = results_post['stage_hb'][stage]
                    
                    stage_rows.append({
                        'Stage': stage,
                        'Pre-Treatment Hours': pre_stage['hrs'],
                        'Pre-Treatment AHI': pre_stage['AHI'],
                        'Pre-Treatment ODI': pre_stage['ODI'],
                        'Pre-Treatment HB': pre_stage['HB'],
                        'Post-Treatment Hours': post_stage['hrs'],
                        'Post-Treatment AHI': post_stage['AHI'],
                        'Post-Treatment ODI': post_stage['ODI'],
                        'Post-Treatment HB': post_stage['HB'],
                        'AHI Change': post_stage['AHI'] - pre_stage['AHI'],
                        'HB Change': post_stage['HB'] - pre_stage['HB'],
                        'HB % Improvement': ((pre_stage['HB'] - post_stage['HB']) / pre_stage['HB'] * 100) if pre_stage['HB'] > 0 else 0
                    })
            
            if stage_rows:
                stage_df = pd.DataFrame(stage_rows)
                stage_df.to_excel(writer, sheet_name='Stage Comparison', index=False)
            
            # Sheet 3: Pre-Treatment Details
            pre_details = {
                'Parameter': ['Filename', 'Duration (hours)', 'AHI', 'ODI', 'Event-Specific HB', 
                             'Global HB', 'Baseline SpO₂', 'Total Events', '95% CI Lower', '95% CI Upper'],
                'Value': [
                    pre_treatment_file.name,
                    f"{results_pre['duration']:.2f}",
                    f"{results_pre['ahi']:.1f}",
                    f"{results_pre['odi']:.1f}",
                    f"{results_pre['total_hb']:.1f}",
                    f"{results_pre.get('global_hb', 0):.1f}",
                    f"{results_pre.get('baseline_used', 'N/A')}",
                    len(results_pre.get('events', [])),
                    f"{results_pre.get('ci', (0, 0))[0]:.1f}",
                    f"{results_pre.get('ci', (0, 0))[1]:.1f}"
                ]
            }
            pd.DataFrame(pre_details).to_excel(writer, sheet_name='Pre-Treatment Details', index=False)
            
            # Sheet 4: Post-Treatment Details
            post_details = {
                'Parameter': ['Filename', 'Duration (hours)', 'AHI', 'ODI', 'Event-Specific HB', 
                             'Global HB', 'Baseline SpO₂', 'Total Events', '95% CI Lower', '95% CI Upper'],
                'Value': [
                    post_treatment_file.name,
                    f"{results_post['duration']:.2f}",
                    f"{results_post['ahi']:.1f}",
                    f"{results_post['odi']:.1f}",
                    f"{results_post['total_hb']:.1f}",
                    f"{results_post.get('global_hb', 0):.1f}",
                    f"{results_post.get('baseline_used', 'N/A')}",
                    len(results_post.get('events', [])),
                    f"{results_post.get('ci', (0, 0))[0]:.1f}",
                    f"{results_post.get('ci', (0, 0))[1]:.1f}"
                ]
            }
            pd.DataFrame(post_details).to_excel(writer, sheet_name='Post-Treatment Details', index=False)
            
            # Sheet 5: Clinical Interpretation
            interpretation_data = {
                'Assessment': ['Treatment Efficacy', 'AHI Response', 'HB Response', 'Recommended Action'],
                'Result': [
                    '✅ Excellent' if (ahi_improvement >= 50 and hb_improvement >= 50) else
                    '✓ Good' if (ahi_improvement >= 30 and hb_improvement >= 30) else
                    '⚠️ Partial' if (ahi_improvement >= 10 and hb_improvement >= 10) else
                    '❌ Poor',
                    f"{ahi_improvement:.1f}% improvement",
                    f"{hb_improvement:.1f}% improvement",
                    'Continue current therapy' if ahi_improvement >= 50 else
                    'Consider therapy optimization' if ahi_improvement >= 30 else
                    'Evaluate alternative treatments'
                ]
            }
            pd.DataFrame(interpretation_data).to_excel(writer, sheet_name='Clinical Interpretation', index=False)
        
        excel_buffer.seek(0)
        
        # Download button for comprehensive Excel report
        st.download_button(
            label="📊 Download Complete Comparison Report (Excel)",
            data=excel_buffer,
            file_name=f"treatment_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.caption("💡 Excel file contains 5 sheets: Summary, Stage Comparison, Pre-Treatment Details, Post-Treatment Details, and Clinical Interpretation")

# =============================================
# BATCH ANALYSIS MODE
# =============================================
st.markdown("---")
st.markdown("### 📦 Batch Mode: Analyze Multiple Files")

batch_files = st.file_uploader(
    "Upload multiple PSG EDF files",
    type=["edf"],
    accept_multiple_files=True,
    key="batch_upload",
    help="Online: ≤5 files, ≤1 GB total. For larger batches, run locally."
)

if batch_files:
    n_files = len(batch_files)
    total_size_gb = sum(f.size for f in batch_files) / 1e9
    
    # Check limits
    if n_files > 5 or total_size_gb > 1.0:
        st.error("⚠️ **Batch Too Large for Online Use**")
        st.markdown(f"""
        Your batch has **{n_files} files** ({total_size_gb:.2f} GB total).
        
        **Online limits:**
        - ≤5 files
        - ≤1 GB total size
        
        **Solution:** Run the app locally (see instructions above) to process larger batches.
        """)
        st.stop()
    
    # Batch settings
    with st.expander("⚙️ Batch Settings", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            batch_include_stages = st.checkbox("Stage-specific results", value=True, key="batch_stages")
            batch_proof_mode = st.selectbox(
                "Proof plots",
                ["None", "Overlay (Azarbarzin-style)", "Full"],
                index=1,
                key="batch_proof"
            )
        
        with col2:
            batch_desat_threshold = st.selectbox(
                "Desaturation threshold",
                ["3%", "4%"],
                index=0,
                key="batch_desat"
            )
            batch_use_global_hb = st.checkbox(
                "Calculate Global HB",
                value=True,
                key="batch_global_hb",
                help="Calculate global hypoxic burden for each file"
            )
    
    # Initialize batch session state
    for key in ['batch_running', 'batch_paused', 'batch_progress', 'batch_results', 'batch_files_processed']:
        if key not in st.session_state:
            if 'progress' in key or 'processed' in key:
                st.session_state[key] = 0
            elif 'running' in key or 'paused' in key:
                st.session_state[key] = False
            else:
                st.session_state[key] = []
    
    # Batch control buttons
    col1, col2, col3 = st.columns(3)
    
    with col1:
        start_btn = st.button(
            "▶️ Run Batch",
            type="primary",
            disabled=st.session_state.batch_running,
            use_container_width=True
        )
    
    with col2:
        pause_btn = st.button(
            "⏸️ Pause",
            disabled=not st.session_state.batch_running or st.session_state.batch_paused,
            use_container_width=True
        )
    
    with col3:
        stop_btn = st.button(
            "⏹️ Stop",
            type="secondary",
            disabled=not st.session_state.batch_running,
            use_container_width=True
        )
    
    # Handle button clicks
    if stop_btn:
        for key in ['batch_running', 'batch_paused', 'batch_progress', 'batch_results', 'batch_files_processed']:
            if 'progress' in key or 'processed' in key:
                st.session_state[key] = 0
            elif 'running' in key or 'paused' in key:
                st.session_state[key] = False
            else:
                st.session_state[key] = []
        st.rerun()
    
    if pause_btn:
        st.session_state.batch_paused = True
        st.session_state.batch_running = False
        st.rerun()
    
    if st.session_state.batch_paused:
        if st.button("▶️ Resume Batch", type="primary", use_container_width=True):
            st.session_state.batch_running = True
            st.session_state.batch_paused = False
            st.rerun()
    
    # Progress indicators
    progress_bar = st.progress(st.session_state.batch_progress)
    status_text = st.empty()
    
    # Run batch processing
    if start_btn or (st.session_state.batch_running and not st.session_state.batch_paused):
        st.session_state.batch_running = True
        start_idx = st.session_state.batch_files_processed
        batch_summary_data = []
        
        desat_thresh_val = 3 if "3%" in batch_desat_threshold else 4
        
        for idx in range(start_idx, n_files):
            if not st.session_state.batch_running:
                break
            
            current_file = batch_files[idx]
            status_text.text(f"📂 Processing {current_file.name} ({idx+1}/{n_files})...")
            progress_bar.progress((idx + 0.1) / n_files)
            
            try:
                # Load file
                raw, temp_path = load_edf_file(current_file, f"temp_batch_{idx}.edf")
                
                if raw is None:
                    st.warning(f"⚠️ Skipping {current_file.name}: Could not load file")
                    continue
                
                # Run analysis
                analyzer = PSGAnalyzer(raw, temp_path)
                
                results = analyzer.run_full_analysis(
                    pre_event_sec=100,
                    desat_start_sec=60,
                    desat_end_sec=120,
                    artifact_filter="Off",
                    desat_threshold=desat_thresh_val,
                    use_global_hb=batch_use_global_hb,
                    preset_baseline=0.0,
                    use_mit_st=False
                )
                
                # Generate PDF (if available)
                if PDF_AVAILABLE:
                    pdf_generator = PDFReportGenerator()
                    buffer = pdf_generator.generate_report(
                        filename=current_file.name,
                        results=results,
                        proof_mode=batch_proof_mode,
                        include_stages=batch_include_stages
                    )
                    # Store results
                    st.session_state.batch_results.append((current_file.name, buffer))
                else:
                    st.warning("⚠️ PDF generation unavailable for batch mode. Skipping PDF creation.")
                
                # Add to summary
                summary_row = {
                    'File': current_file.name,
                    'Duration (h)': f"{results['duration']:.1f}",
                    'AHI': f"{results['ahi']:.1f}",
                    'ODI': f"{results['odi']:.1f}",
                    'Obstructive HB': f"{results['total_hb']:.2f}"
                }
                
                if results.get('global_hb') is not None:
                    summary_row['Global HB'] = f"{results['global_hb']:.2f}"
                    summary_row['Baseline'] = f"{results['baseline_used']:.1f}%"
                
                batch_summary_data.append(summary_row)
                
                # Cleanup
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
            except Exception as e:
                st.error(f"❌ Error processing {current_file.name}: {str(e)}")
                continue
            
            # Update progress
            st.session_state.batch_files_processed = idx + 1
            progress_bar.progress((idx + 1) / n_files)
        
        # Generate master summary and ZIP
        status_text.text("📊 Generating master summary...")
        
        if PDF_AVAILABLE and len(st.session_state.batch_results) > 0:
            pdf_generator = PDFReportGenerator()
            master_buffer = pdf_generator.generate_batch_summary(batch_summary_data)
        
            # Create ZIP file
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Add individual reports
                for filename, pdf_buffer in st.session_state.batch_results:
                    zf.writestr(
                        f"Reports/HB_Report_{filename.replace('.edf', '')}.pdf",
                        pdf_buffer.getvalue()
                    )
                
                # Add master summary
            zf.writestr("Master_Summary.pdf", master_buffer.getvalue())
        
        zip_buffer.seek(0)
        
        # Success message
        progress_bar.progress(1.0)
        status_text.text("✅ Batch processing complete!")
        st.success(f"**Batch Complete!** Generated {len(st.session_state.batch_results)} reports.")
        
        # Download button
        st.download_button(
            label=f"⬇️ Download All Reports ({len(st.session_state.batch_results)} files)",
            data=zip_buffer.getvalue(),
            file_name=f"HB_Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True
        )
        
        # Reset batch state
        for key in ['batch_running', 'batch_paused', 'batch_progress', 'batch_files_processed']:
            if 'progress' in key or 'processed' in key:
                st.session_state[key] = 0
            else:
                st.session_state[key] = False
        
        st.session_state.batch_results = []

# --------------------------------------------------------------
# FOOTER
# --------------------------------------------------------------
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p><strong>Hypoxic Burden Calculator</strong> — Open Source Sleep Apnea Analysis</p>
    <p>
        🐙 <a href="https://github.com/Apolloplectic/hypoxic-burden-edf">GitHub</a> • 
        📄 <a href="https://doi.org/10.5281/zenodo.17561726">DOI: 10.5281/zenodo.17561726</a> • 
        📧 <a href="mailto:sam.johnson9797@gmail.com">Contact</a>
    </p>
    <p><small>Built with Streamlit • MNE • {yasa_status} • WFDB</small></p>
    <p><small>Cite: Azarbarzin A, et al. <em>Eur Heart J</em> 2019;40:1149-1157</small></p>
</div>
""".format(yasa_status="YASA ✅" if YASA_AVAILABLE else "YASA ❌"), unsafe_allow_html=True)
